"""
Salesforce Picklist Export Script - Enhanced Version
Exports picklist metadata for multiple objects to Excel format
Includes comprehensive statistics and error tracking

SETUP:
1. Install Python if it is not installed
2. Install required packages:
   pip install simple-salesforce openpyxl requests

3. Update credentials in TWO WAYS:

   METHOD 1 - Direct in script (lines 38-41):
   SF_USERNAME = 'your_username@example.com'
   SF_PASSWORD = 'your_password'
   SF_SECURITY_TOKEN = 'your_security_token'
   SF_DOMAIN = 'login'

   METHOD 2 - Environment Variables (recommended for security):
   export SF_USERNAME='your_username@example.com'
   export SF_PASSWORD='your_password'
   export SF_SECURITY_TOKEN='your_security_token'
   export SF_DOMAIN='login'

4. for this line export SF_DOMAIN='login'
    login-->for Production org
    test-->for Sandbox org

4. Run:
   python picklist_export.py
"""

import os
import sys
import time
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import requests
from simple_salesforce import Salesforce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ===========================================
# CONFIGURATION - UPDATE CREDENTIALS HERE
# ===========================================

SF_USERNAME = os.getenv('SF_USERNAME', 'nahidhasan00619436@agentforce.com')
SF_PASSWORD = os.getenv('SF_PASSWORD', 'nahid@gfgg2')
SF_SECURITY_TOKEN = os.getenv('SF_SECURITY_TOKEN', 'YaeOEq6Rw1XUMbPJ5EY3etclw')
SF_DOMAIN = os.getenv('SF_DOMAIN', 'login')

# Objects to export
OBJECTS_TO_EXPORT = [
'Account'
]

API_VERSION = '64.0'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, f'Picklist_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
DEBUG_MODE = False


# ===========================================
# HELPER CLASSES
# ===========================================

class FieldInfo:
    """Represents picklist field metadata"""
    def __init__(self, api_name: str, label: str):
        self.api_name = api_name
        self.label = label


class PicklistValueDetail:
    """Represents a single picklist value"""
    def __init__(self, label: str, value: str, is_active: bool = True):
        self.label = label
        self.value = value
        self.is_active = is_active


class ProcessingResult:
    """Stores processing results for an object"""
    def __init__(self):
        self.values_processed = 0
        self.inactive_values = 0
        self.rows: List[List[str]] = []
        self.picklist_fields_count = 0
        self.object_exists = True
        self.error_message = None


# ===========================================
# MAIN EXPORT CLASS
# ===========================================

class PicklistExporter:
    """Main exporter class with enhanced statistics"""
    
    def __init__(self, username: str, password: str, security_token: str, domain: str = 'login'):
        """Initialize Salesforce connection"""
        print("=== Initializing Salesforce Connection ===")
        try:
            self.sf = Salesforce(
                username=username,
                password=password,
                security_token=security_token,
                domain=domain
            )
            self.base_url = f"https://{self.sf.sf_instance}"
            self.session_id = self.sf.session_id
            self.headers = {
                'Authorization': f'Bearer {self.session_id}',
                'Content-Type': 'application/json'
            }
            print(f"✅ Connected to: {self.base_url}")
            print(f"✅ User: {username}")
            print()
        except Exception as e:
            print(f"❌ Connection failed: {str(e)}")
            sys.exit(1)
    
    def export_picklists(self, object_names: List[str]) -> Tuple[str, Dict]:
        """
        Main export method with enhanced statistics tracking
        Returns: (output_path, statistics_dict)
        """
        print("=== Starting Picklist Export ===")
        print(f"Total objects to process: {len(object_names)}")
        print()
        
        # Enhanced statistics tracking
        stats = {
            'total_objects': len(object_names),
            'successful_objects': 0,
            'failed_objects': 0,
            'objects_not_found': 0,
            'objects_with_zero_picklists': 0,
            'objects_with_picklists': 0,
            'total_picklist_fields': 0,
            'total_values': 0,
            'total_active_values': 0,
            'total_inactive_values': 0,
            'failed_object_details': [],  # List of {name, reason}
            'objects_without_picklists': [],
            'objects_not_found_list': []
        }
        
        all_rows = []
        header = ['Object', 'Field Label', 'Field API', 'Picklist Value Label', 'Picklist Value API', 'Status']
        all_rows.append(header)
        
        # Process each object
        for i, obj_name in enumerate(object_names, 1):
            print(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            
            try:
                result = self._process_object(obj_name)
                
                # Track different scenarios
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({
                        'name': obj_name,
                        'reason': 'Object does not exist in org'
                    })
                    print(f"  ⚠️  Object not found in org")
                    
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    print(f"  ℹ️  No picklist fields found")
                    
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    all_rows.extend(result.rows)
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    
                    print(f"  ✅ Fields: {result.picklist_fields_count}, Active: {result.values_processed - result.inactive_values}, Inactive: {result.inactive_values}")
                
            except Exception as e:
                error_msg = str(e)
                print(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
                continue
            
            print()
        
        # Create Excel file
        print("=== Creating Excel File ===")
        output_path = self._create_excel_file(all_rows)
        
        return output_path, stats
    
    def _process_object(self, obj_name: str) -> ProcessingResult:
        """Process a single object"""
        result = ProcessingResult()
        
        # Check if object exists
        try:
            getattr(self.sf, obj_name).describe()
        except Exception as e:
            if 'NOT_FOUND' in str(e) or 'INVALID_TYPE' in str(e):
                result.object_exists = False
                return result
            raise
        
        # Get picklist fields
        picklist_fields = self._get_picklist_fields(obj_name)
        result.picklist_fields_count = len(picklist_fields)
        
        if not picklist_fields:
            return result
        
        print(f"  Found {len(picklist_fields)} picklist fields")
        
        # Resolve EntityDefinition Id
        entity_def_id = self._resolve_entity_definition_id(obj_name)
        if entity_def_id:
            print(f"  EntityDefinition.Id: {entity_def_id}")
        
        # Process each field
        for field_api, field_info in picklist_fields.items():
            values = self._query_picklist_values_with_fallback(
                obj_name, entity_def_id, field_api
            )
            
            if not values:
                print(f"    Field: {field_api} - No values found")
                continue
            
            print(f"    Field: {field_api} - {len(values)} values")
            
            if DEBUG_MODE and values:
                print(f"      [DEBUG] Sample values:")
                for idx, v in enumerate(values[:3]):
                    print(f"        {idx+1}. {v.label} = {v.value} (Active: {v.is_active})")
            
            # Build rows
            for value in values:
                is_active = value.is_active if value.is_active is not None else True
                status = 'Active' if is_active else 'Inactive'
                
                if not is_active:
                    result.inactive_values += 1
                
                row = [
                    obj_name,
                    field_info.label,
                    field_api,
                    value.label,
                    value.value,
                    status
                ]
                result.rows.append(row)
                result.values_processed += 1
        
        return result
    
    def _get_picklist_fields(self, object_name: str) -> Dict[str, FieldInfo]:
        """Get all picklist fields"""
        fields_dict = {}
        
        try:
            obj_describe = getattr(self.sf, object_name).describe()
            
            for field in obj_describe['fields']:
                field_type = field['type']
                
                if field_type in ['picklist', 'multipicklist']:
                    field_info = FieldInfo(
                        api_name=field['name'],
                        label=field['label']
                    )
                    fields_dict[field_info.api_name] = field_info
        
        except Exception as e:
            print(f"  ERROR in _get_picklist_fields: {str(e)}")
        
        return fields_dict
    
    def _resolve_entity_definition_id(self, object_name: str) -> Optional[str]:
        """Resolve EntityDefinition Id"""
        try:
            query = f"SELECT Id FROM EntityDefinition WHERE QualifiedApiName = '{object_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            params = {'q': query}
            
            response = requests.get(url, headers=self.headers, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                records = data.get('records', [])
                if records:
                    return records[0]['Id']
        
        except Exception as e:
            print(f"  ERROR resolveEntityDefinitionId: {str(e)}")
        
        return None
    
    def _query_picklist_values_with_fallback(
        self, 
        object_name: str, 
        entity_def_id: Optional[str], 
        field_name: str
    ) -> List[PicklistValueDetail]:
        """Try multiple approaches to get picklist values"""
        
        values = self._query_field_definition_tooling(object_name, field_name)
        if values:
            return values
        
        if entity_def_id:
            values = self._query_custom_field_tooling(entity_def_id, field_name)
            if values:
                return values
        
        values = self._query_custom_field_tooling_table_enum(object_name, field_name)
        if values:
            return values
        
        values = self._query_rest_describe_for_picklist(object_name, field_name)
        if values:
            return values
        
        return []
    
    def _query_field_definition_tooling(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Fallback method 1"""
        try:
            query = f"SELECT Metadata FROM FieldDefinition WHERE EntityDefinition.QualifiedApiName = '{object_name}' AND QualifiedApiName = '{field_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            params = {'q': query}
            
            response = requests.get(url, headers=self.headers, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                records = data.get('records', [])
                if records:
                    metadata = records[0].get('Metadata', {})
                    
                    if DEBUG_MODE:
                        print(f"      [DEBUG] FieldDefinition metadata for {field_name}:")
                        import json
                        print(json.dumps(metadata.get('valueSet', {}), indent=2)[:500])
                    
                    return self._parse_value_set(metadata)
        
        except Exception as e:
            print(f"      ERROR queryFieldDefinitionTooling: {str(e)}")
        
        return []
    
    def _query_custom_field_tooling(self, entity_def_id: str, field_name: str) -> List[PicklistValueDetail]:
        """Fallback method 2"""
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{entity_def_id}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            params = {'q': query}
            
            response = requests.get(url, headers=self.headers, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                records = data.get('records', [])
                if records:
                    metadata = records[0].get('Metadata', {})
                    return self._parse_value_set(metadata)
        
        except Exception as e:
            print(f"      ERROR queryCustomFieldTooling: {str(e)}")
        
        return []
    
    def _query_custom_field_tooling_table_enum(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Fallback method 3"""
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{object_name}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            params = {'q': query}
            
            response = requests.get(url, headers=self.headers, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                records = data.get('records', [])
                if records:
                    metadata = records[0].get('Metadata', {})
                    return self._parse_value_set(metadata)
        
        except Exception as e:
            print(f"      ERROR queryCustomFieldToolingTableEnum: {str(e)}")
        
        return []
    
    def _query_rest_describe_for_picklist(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Fallback method 4"""
        try:
            url = f"{self.base_url}/services/data/v{API_VERSION}/sobjects/{object_name}/describe"
            response = requests.get(url, headers=self.headers, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                fields = data.get('fields', [])
                
                for field in fields:
                    if field['name'].lower() == field_name.lower():
                        picklist_values = field.get('picklistValues', [])
                        results = []
                        for pv in picklist_values:
                            active_value = pv.get('active')
                            is_active = True if active_value is None else bool(active_value)
                            
                            detail = PicklistValueDetail(
                                label=pv.get('label', ''),
                                value=pv.get('value', ''),
                                is_active=is_active
                            )
                            results.append(detail)
                        return results
        
        except Exception as e:
            print(f"      ERROR queryRestDescribeForPicklist: {str(e)}")
        
        return []
    
    def _parse_value_set(self, metadata: dict) -> List[PicklistValueDetail]:
        """Parse valueSet from metadata"""
        results = []
        
        try:
            value_set = metadata.get('valueSet', {})
            if not value_set:
                return results
            
            if 'valueSetDefinition' in value_set:
                vsd = value_set['valueSetDefinition']
                values = vsd.get('value', [])
            elif 'value' in value_set:
                values = value_set['value']
            else:
                return results
            
            for v in values:
                label = v.get('label', '')
                value_name = v.get('valueName') or v.get('value', '')
                
                is_active_raw = v.get('isActive')
                if is_active_raw is None:
                    is_active = True
                else:
                    is_active = bool(is_active_raw)
                
                detail = PicklistValueDetail(
                    label=label,
                    value=value_name,
                    is_active=is_active
                )
                results.append(detail)
        
        except Exception as e:
            print(f"      ERROR parseValueSet: {str(e)}")
        
        return results
    
    def _create_excel_file(self, rows: List[List[str]]) -> str:
        """Create Excel file with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Picklist Export"
        
        for row in rows:
            ws.append(row)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = "A2"
        
        wb.save(OUTPUT_FILE)
        print(f"✅ Excel file created: {OUTPUT_FILE}")
        print(f"✅ File location: {SCRIPT_DIR}")
        print(f"✅ Total data rows: {len(rows) - 1}")
        
        return OUTPUT_FILE


# ===========================================
# MAIN EXECUTION
# ===========================================

def format_runtime(seconds: float) -> str:
    """Format runtime in HH:MM:SS format"""
    td = timedelta(seconds=int(seconds))
    hours, remainder = divmod(td.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def print_statistics(stats: Dict, runtime_formatted: str, output_file: str):
    """Print comprehensive statistics - ONLY CALLED ONCE"""
    print("=" * 70)
    print("✅ EXPORT COMPLETED SUCCESSFULLY!")
    print("=" * 70)
    print()
    
    # Runtime
    print("⏱️  EXECUTION TIME:")
    print("-" * 70)
    print(f"Total Runtime: {runtime_formatted}")
    print()
    
    # Object Processing Summary
    print("📦 OBJECT PROCESSING SUMMARY:")
    print("-" * 70)
    print(f"Total Objects in List:          {stats['total_objects']}")
    print(f"✅ Successfully Processed:       {stats['successful_objects']}")
    print(f"❌ Failed to Process:            {stats['failed_objects']}")
    print(f"⚠️  Objects Not Found in Org:    {stats['objects_not_found']}")
    print(f"ℹ️  Objects with Zero Picklists:  {stats['objects_with_zero_picklists']}")
    print(f"📋 Objects with Picklists:       {stats['objects_with_picklists']}")
    print()
    
    # Picklist Fields & Values
    print("📊 PICKLIST STATISTICS:")
    print("-" * 70)
    print(f"Total Picklist Fields:          {stats['total_picklist_fields']}")
    print(f"Total Picklist Values:          {stats['total_values']}")
    print(f"✅ Active Values:                {stats['total_active_values']}")
    print(f"❌ Inactive Values:              {stats['total_inactive_values']}")
    
    if stats['total_values'] > 0:
        active_pct = (stats['total_active_values'] / stats['total_values']) * 100
        inactive_pct = (stats['total_inactive_values'] / stats['total_values']) * 100
        print(f"Active Percentage:              {active_pct:.1f}%")
        print(f"Inactive Percentage:            {inactive_pct:.1f}%")
    print()
    
    # Details for Objects Not Found
    if stats['objects_not_found'] > 0:
        print("⚠️  OBJECTS NOT FOUND IN ORG:")
        print("-" * 70)
        for obj in stats['objects_not_found_list']:
            print(f"   • {obj}")
        print()
    
    # Details for Objects Without Picklists
    if stats['objects_with_zero_picklists'] > 0:
        print("ℹ️  OBJECTS WITHOUT PICKLIST FIELDS:")
        print("-" * 70)
        for obj in stats['objects_without_picklists']:
            print(f"   • {obj}")
        print()
    
    # Details for Failed Objects
    if stats['failed_objects'] > 0:
        print("❌ FAILED OBJECTS (WITH REASONS):")
        print("-" * 70)
        for detail in stats['failed_object_details']:
            print(f"   • {detail['name']}")
            print(f"     Reason: {detail['reason']}")
        print()
    
    # Output File
    print("📄 OUTPUT FILE:")
    print("-" * 70)
    print(f"Location: {output_file}")
    print("=" * 70)


def main():
    """Main execution function"""
    start_time = time.time()
    
    print("=" * 70)
    print("SALESFORCE PICKLIST EXPORT TO EXCEL")
    print("=" * 70)
    print()
    
    using_env = 'SF_USERNAME' in os.environ
    if using_env:
        print("🔐 Using credentials from ENVIRONMENT VARIABLES")
    else:
        print("🔐 Using credentials from SCRIPT")
    print()
    
    # Validate credentials
    if not SF_USERNAME or SF_USERNAME == 'your_username@example.com':
        print("❌ ERROR: Please provide credentials")
        print()
        print("METHOD 1 - Update in script (lines 48-51):")
        print("  SF_USERNAME = 'your_username@example.com'")
        print("  SF_PASSWORD = 'your_password'")
        print("  SF_SECURITY_TOKEN = 'your_security_token'")
        print()
        print("METHOD 2 - Use environment variables (recommended):")
        print("  export SF_USERNAME='your_username@example.com'")
        print("  export SF_PASSWORD='your_password'")
        print("  export SF_SECURITY_TOKEN='your_security_token'")
        print()
        sys.exit(1)
    
    if not SF_PASSWORD or SF_PASSWORD == 'your_password':
        print("❌ ERROR: Please update SF_PASSWORD")
        sys.exit(1)
    
    if not SF_SECURITY_TOKEN or SF_SECURITY_TOKEN == 'your_security_token':
        print("❌ ERROR: Please update SF_SECURITY_TOKEN")
        sys.exit(1)
    
    print(f"📁 Excel file will be saved in: {SCRIPT_DIR}")
    print()
    
    try:
        exporter = PicklistExporter(
            username=SF_USERNAME,
            password=SF_PASSWORD,
            security_token=SF_SECURITY_TOKEN,
            domain=SF_DOMAIN
        )
        
        output_file, stats = exporter.export_picklists(OBJECTS_TO_EXPORT)
        
        # Calculate runtime ONCE
        end_time = time.time()
        runtime_seconds = end_time - start_time
        runtime_formatted = format_runtime(runtime_seconds)
        
        # Print statistics ONLY ONCE
        print_statistics(stats, runtime_formatted, output_file)
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Export cancelled by user")
        end_time = time.time()
        runtime_seconds = end_time - start_time
        runtime_formatted = format_runtime(runtime_seconds)
        print(f"⏱️  Runtime before cancellation: {runtime_formatted}")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        end_time = time.time()
        runtime_seconds = end_time - start_time
        runtime_formatted = format_runtime(runtime_seconds)
        print(f"\n⏱️  Runtime before failure: {runtime_formatted}")
        sys.exit(1)


if __name__ == "__main__":
    main()