'''
Salesforce Picklist Export - Optimized GUI Version
Fixed: UI freezing, threading issues, responsiveness
Optimized: Search performance, list updates, memory usage

Prerequisites (Installation):
   pip install simple-salesforce openpyxl requests customtkinter
'''
import os
import sys
import time
import requests
import tkinter as tk 
import threading
import queue
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple, Set

# Third-party libraries
from simple_salesforce import Salesforce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# GUI Library
import customtkinter as ctk
from tkinter import messagebox, filedialog, END

# Set appearance mode and default color theme
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

# ===========================================
# CONFIGURATION / GLOBAL CONSTANTS
# ===========================================

API_VERSION = '65.0'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ===========================================
# HELPER CLASSES 
# ===========================================

class FieldInfo:
    """Represents picklist field metadata"""
    def __init__(self, api_name: str, label: str):
        self.api_name = api_name
        self.label = label


class PicklistValueDetail:
    """Represents a single picklist value"""
    def __init__(self, label: str, value: str, is_active: bool = True):
        self.label = label
        self.value = value
        self.is_active = is_active


class ProcessingResult:
    """Stores processing results for an object"""
    def __init__(self):
        self.values_processed = 0
        self.inactive_values = 0
        self.rows: List[List[str]] = []
        self.picklist_fields_count = 0
        self.object_exists = True
        self.error_message = None


# ===========================================
# MAIN EXPORT CLASS 
# ===========================================

class PicklistExporter:
    """Main exporter class with enhanced statistics"""
    
    def __init__(self, username: str, password: str, security_token: str, domain: str = 'login', status_callback=None):
        """Initialize Salesforce connection"""
        self.status_callback = status_callback
        self.all_org_objects: List[str] = []
        self._cancel_flag = False
        
        if self.status_callback:
            self.status_callback("Initializing Salesforce Connection...")
            
        try:
            self.sf = Salesforce(
                username=username,
                password=password,
                security_token=security_token,
                domain=domain
            )
            self.base_url = f"https://{self.sf.sf_instance}"
            self.session_id = self.sf.session_id
            self.headers = {
                'Authorization': f'Bearer {self.session_id}',
                'Content-Type': 'application/json'
            }
            if self.status_callback:
                self.status_callback(f"✅ Connected to: {self.base_url}")
            
            self._fetch_all_org_objects()

        except Exception as e:
            if self.status_callback:
                self.status_callback(f"❌ Connection failed: {str(e)}")
            raise

    def _fetch_all_org_objects(self):
        """Fetches all SObjects (Standard and Custom) from the org"""
        self._log_status("Fetching all available SObjects from the organization...")
        try:
            response = self.sf.describe()
            self.all_org_objects = sorted([
                obj['name'] for obj in response['sobjects'] 
                if obj.get('queryable', False) and not obj.get('deprecatedAndHidden', False)
            ])
            self._log_status(f"✅ Found {len(self.all_org_objects)} queryable objects.")
        except Exception as e:
            self._log_status(f"❌ Failed to fetch all SObjects: {str(e)}")
            self.all_org_objects = []
            
    def get_all_objects(self) -> List[str]:
        """Accessor for the fetched object list"""
        return self.all_org_objects

    def cancel_export(self):
        """Set cancellation flag"""
        self._cancel_flag = True
        self._log_status("⚠️ Export cancellation requested...")

    def _log_status(self, message: str):
        """Internal helper to send log messages back to the GUI"""
        if self.status_callback:
            self.status_callback(message, verbose=True) 

    def export_picklists(self, object_names: List[str], output_path: str) -> Tuple[str, Dict]:
        """Main export method with cancellation support"""
        self._cancel_flag = False
        self._log_status("=== Starting Picklist Export ===")
        self._log_status(f"Total objects to process: {len(object_names)}")
        
        stats = {
            'total_objects': len(object_names), 'successful_objects': 0, 'failed_objects': 0, 'objects_not_found': 0,
            'objects_with_zero_picklists': 0, 'objects_with_picklists': 0, 'total_picklist_fields': 0, 'total_values': 0,
            'total_active_values': 0, 'total_inactive_values': 0, 'failed_object_details': [],
            'objects_without_picklists': [], 'objects_not_found_list': [], 'cancelled': False
        }
        
        all_rows = [['Object', 'Field Label', 'Field API', 'Picklist Value Label', 'Picklist Value API', 'Status']]
        
        for i, obj_name in enumerate(object_names, 1):
            # Check cancellation flag
            if self._cancel_flag:
                self._log_status("⚠️ Export cancelled by user")
                stats['cancelled'] = True
                break
                
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({'name': obj_name, 'reason': 'Object does not exist in org'})
                    self._log_status(f"  ⚠️  Object not found in org")
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  ℹ️  No picklist fields found")
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    all_rows.extend(result.rows)
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    self._log_status(f"  ✅ Fields: {result.picklist_fields_count}, Active: {result.values_processed - result.inactive_values}, Inactive: {result.inactive_values}")
            except Exception as e:
                if self._cancel_flag:
                    break
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({'name': obj_name, 'reason': error_msg})
            self._log_status("")
        
        if not self._cancel_flag:
            self._log_status("=== Creating Excel File ===")
            final_output_path = self._create_excel_file(all_rows, output_path)
            return final_output_path, stats
        else:
            return None, stats
    
    def _process_object(self, obj_name: str) -> ProcessingResult:
        result = ProcessingResult()
        
        if self._cancel_flag:
            return result
            
        try:
            getattr(self.sf, obj_name).describe()
        except Exception as e:
            if 'NOT_FOUND' in str(e) or 'INVALID_TYPE' in str(e):
                result.object_exists = False
                return result
            raise
        
        picklist_fields = self._get_picklist_fields(obj_name)
        result.picklist_fields_count = len(picklist_fields)
        if not picklist_fields: return result
        self._log_status(f"  Found {len(picklist_fields)} picklist fields")
        entity_def_id = self._resolve_entity_definition_id(obj_name)
        if entity_def_id: self._log_status(f"  EntityDefinition.Id: {entity_def_id}")
        
        for field_api, field_info in picklist_fields.items():
            if self._cancel_flag:
                break
            values = self._query_picklist_values_with_fallback(obj_name, entity_def_id, field_api)
            if not values: continue
            self._log_status(f"    Field: {field_api} - {len(values)} values")
            for value in values:
                is_active = value.is_active if value.is_active is not None else True
                status = 'Active' if is_active else 'Inactive'
                if not is_active: result.inactive_values += 1
                row = [obj_name, field_info.label, field_api, value.label, value.value, status]
                result.rows.append(row)
                result.values_processed += 1
        return result
    
    def _get_picklist_fields(self, object_name: str) -> Dict[str, FieldInfo]:
        fields_dict = {}
        try:
            obj_describe = getattr(self.sf, object_name).describe()
            for field in obj_describe['fields']:
                if field['type'] in ['picklist', 'multipicklist']:
                    fields_dict[field['name']] = FieldInfo(api_name=field['name'], label=field['label'])
        except Exception as e:
            self._log_status(f"  ERROR in _get_picklist_fields: {str(e)}")
        return fields_dict
    
    def _resolve_entity_definition_id(self, object_name: str) -> Optional[str]:
        try:
            query = f"SELECT Id FROM EntityDefinition WHERE QualifiedApiName = '{object_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records: return records[0]['Id']
        except Exception as e:
            self._log_status(f"  ERROR resolveEntityDefinitionId: {str(e)}")
        return None
    
    def _query_picklist_values_with_fallback(self, object_name: str, entity_def_id: Optional[str], field_name: str) -> List[PicklistValueDetail]:
        values = self._query_field_definition_tooling(object_name, field_name)
        if values: return values
        if entity_def_id:
            values = self._query_custom_field_tooling(entity_def_id, field_name)
            if values: return values
        values = self._query_custom_field_tooling_table_enum(object_name, field_name)
        if values: return values
        values = self._query_rest_describe_for_picklist(object_name, field_name)
        if values: return values
        return []
    
    def _query_field_definition_tooling(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        try:
            query = f"SELECT Metadata FROM FieldDefinition WHERE EntityDefinition.QualifiedApiName = '{object_name}' AND QualifiedApiName = '{field_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records: return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryFieldDefinitionTooling: {str(e)}")
        return []
    
    def _query_custom_field_tooling(self, entity_def_id: str, field_name: str) -> List[PicklistValueDetail]:
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{entity_def_id}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records: return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryCustomFieldTooling: {str(e)}")
        return []
    
    def _query_custom_field_tooling_table_enum(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{object_name}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records: return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryCustomFieldToolingTableEnum: {str(e)}")
        return []
    
    def _query_rest_describe_for_picklist(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        try:
            url = f"{self.base_url}/services/data/v{API_VERSION}/sobjects/{object_name}/describe"
            response = requests.get(url, headers=self.headers, timeout=60)
            if response.status_code == 200:
                for field in response.json().get('fields', []):
                    if field['name'].lower() == field_name.lower():
                        results = []
                        for pv in field.get('picklistValues', []):
                            # CRITICAL FIX: Handle active field properly
                            active_value = pv.get('active')
                            # If 'active' key exists, use its boolean value
                            # If 'active' key is missing, default to True
                            is_active = True if active_value is None else bool(active_value)
                            
                            results.append(PicklistValueDetail(
                                label=pv.get('label', ''), 
                                value=pv.get('value', ''), 
                                is_active=is_active
                            ))
                        return results
        except Exception as e:
            self._log_status(f"      ERROR queryRestDescribeForPicklist: {str(e)}")
        return []
    
    def _parse_value_set(self, metadata: dict) -> List[PicklistValueDetail]:
        results = []
        try:
            value_set = metadata.get('valueSet', {})
            if not value_set: return results
            values = value_set.get('valueSetDefinition', {}).get('value', []) or value_set.get('value', [])
            for v in values:
                # CRITICAL FIX: Handle isActive field properly
                is_active_raw = v.get('isActive')
                # If isActive is None (key missing), default to True
                # If isActive exists, convert to boolean
                if is_active_raw is None:
                    is_active = True
                else:
                    is_active = bool(is_active_raw)
                
                results.append(PicklistValueDetail(
                    label=v.get('label', ''), 
                    value=v.get('valueName') or v.get('value', ''), 
                    is_active=is_active
                ))
        except Exception as e:
            self._log_status(f"      ERROR parseValueSet: {str(e)}")
        return results
    
    def _create_excel_file(self, rows: List[List[str]], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Picklist Export"
        for row in rows: ws.append(row)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        ws.freeze_panes = "A2"
        wb.save(output_path)
        self._log_status(f"✅ Excel file created: {output_path}")
        self._log_status(f"✅ Total data rows: {len(rows) - 1}")
        return output_path


# ===========================================
# GUI HELPER FUNCTIONS
# ===========================================

def format_runtime(seconds: float) -> str:
    """Format runtime in HH:MM:SS format"""
    td = timedelta(seconds=int(seconds))
    hours, remainder = divmod(td.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def print_statistics(stats: Dict, runtime_formatted: str, output_file: str):
    """Prints comprehensive statistics to the console"""
    print("\n" + "=" * 70)
    if stats.get('cancelled', False):
        print("⚠️ EXPORT CANCELLED BY USER")
    else:
        print("✅ EXPORT COMPLETED SUCCESSFULLY! (Statistics Detail)")
    print("=" * 70)
    print(f"Total Runtime: {runtime_formatted}")
    print(f"Total Objects in List:          {stats['total_objects']}")
    print(f"✅ Successfully Processed:       {stats['successful_objects']}")
    print(f"❌ Failed to Process:            {stats['failed_objects']}")
    print(f"⚠️  Objects Not Found in Org:    {stats['objects_not_found']}")
    print(f"Total Picklist Fields:          {stats['total_picklist_fields']}")
    print(f"Total Picklist Values:          {stats['total_values']}")
    print(f"✅ Active Values:                {stats['total_active_values']}")
    print(f"❌ Inactive Values:              {stats['total_inactive_values']}")
    if output_file:
        print(f"Output File: {output_file}")
    if stats['failed_objects'] > 0:
        print("\n❌ FAILED OBJECTS (REASONS):")
        for detail in stats['failed_object_details']:
            print(f"   • {detail['name']}: {detail['reason']}")
    print("=" * 70)


# ===========================================
# GUI IMPLEMENTATION
# ===========================================

class PicklistExportGUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("Salesforce Picklist Exporter")
        self.geometry("1200x768")
        self.minsize(1000, 600)
        
        self.sf_exporter: Optional[PicklistExporter] = None
        self.all_org_objects: List[str] = []
        self.selected_objects: Set[str] = set()
        self.export_thread: Optional[threading.Thread] = None
        self.message_queue = queue.Queue()
        self.is_exporting = False
        
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        self.login_frame = ctk.CTkFrame(self)
        self.export_frame = ctk.CTkFrame(self)
        
        self.login_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        self.export_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        
        self._setup_login_frame()
        self._setup_export_frame()
        
        self.export_frame.grid_forget()
        
        # Start message queue processor
        self._process_message_queue()

    def _process_message_queue(self):
        """Process messages from background thread safely"""
        try:
            while True:
                msg_type, data = self.message_queue.get_nowait()
                
                if msg_type == "status":
                    message, verbose = data
                    self._update_status_internal(message, verbose)
                elif msg_type == "export_complete":
                    output_path, stats, runtime = data
                    self._handle_export_complete(output_path, stats, runtime)
                elif msg_type == "export_error":
                    error_msg = data
                    self._handle_export_error(error_msg)
                    
        except queue.Empty:
            pass
        finally:
            # Schedule next check
            self.after(100, self._process_message_queue)

    # ==================================
    # Screen 1: Login & Authentication
    # ==================================
    
    def _setup_login_frame(self):
        login_frame = self.login_frame
        login_frame.columnconfigure(1, weight=1)
        
        ctk.CTkLabel(login_frame, text="Salesforce Login", font=ctk.CTkFont(size=30, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(50, 40))

        def create_input_row(parent, row, label_text, password_mode=False):
            ctk.CTkLabel(parent, text=label_text, anchor="w", font=ctk.CTkFont(size=14)).grid(row=row, column=0, padx=10, pady=15, sticky="w")
            entry = ctk.CTkEntry(parent, width=350, show="*" if password_mode else "")
            entry.grid(row=row, column=1, padx=10, pady=15, sticky="ew")
            return entry

        self.username_entry = create_input_row(login_frame, 1, "Username:")
        self.password_entry = create_input_row(login_frame, 2, "Password:", password_mode=True)
        self.token_entry = create_input_row(login_frame, 3, "Security Token:", password_mode=True)

        ctk.CTkLabel(login_frame, text="Org Type:", anchor="w", font=ctk.CTkFont(size=14)).grid(row=4, column=0, padx=10, pady=15, sticky="w")
        self.org_type_var = ctk.StringVar(value="Production")
        radio_prod = ctk.CTkRadioButton(login_frame, text="Production", variable=self.org_type_var, value="Production")
        radio_test = ctk.CTkRadioButton(login_frame, text="Sandbox/Test", variable=self.org_type_var, value="Sandbox")
        
        radio_prod.grid(row=4, column=1, padx=(10, 5), pady=15, sticky="w")
        radio_test.grid(row=4, column=1, padx=(140, 10), pady=15, sticky="w")
        
        self.login_button = ctk.CTkButton(login_frame, text="Login to Salesforce", command=self.login_action, height=50, font=ctk.CTkFont(size=16, weight="bold"))
        self.login_button.grid(row=5, column=0, columnspan=2, pady=50, sticky="ew", padx=10)

    def login_action(self):
        """Login with threading to prevent UI freeze"""
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        token = self.token_entry.get().strip()
        domain = 'test' if self.org_type_var.get() == 'Sandbox' else 'login'
        
        if not all([username, password, token]):
            messagebox.showerror("Input Error", "All fields (Username, Password, Security Token) are required.")
            return

        self.login_button.configure(state="disabled", text="Connecting...")
        
        def login_worker():
            try:
                exporter = PicklistExporter(
                    username=username, 
                    password=password, 
                    security_token=token, 
                    domain=domain,
                    status_callback=self.queue_status_update
                )
                
                # Success - update UI in main thread
                self.after(0, lambda: self._handle_login_success(exporter))
                
            except Exception as e:
                # Error - update UI in main thread
                self.after(0, lambda: self._handle_login_error(str(e)))
        
        # Start login in background thread
        threading.Thread(target=login_worker, daemon=True).start()

    def _handle_login_success(self, exporter):
        """Handle successful login (called in main thread)"""
        self.sf_exporter = exporter
        messagebox.showinfo("Success", "Successfully connected to Salesforce!")
        
        self.all_org_objects = self.sf_exporter.get_all_objects()
        
        # Switch to Export Frame
        self.login_frame.grid_forget()
        self.export_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        self.populate_available_objects(self.all_org_objects)
        self.populate_selected_objects()

    def _handle_login_error(self, error_msg):
        """Handle login error (called in main thread)"""
        messagebox.showerror("Login Failed", f"Connection Error: {error_msg}")
        self.sf_exporter = None
        self.login_button.configure(state="normal", text="Login to Salesforce")

    def queue_status_update(self, message: str, verbose: bool = False):
        """Queue status update from background thread"""
        self.message_queue.put(("status", (message, verbose)))

    def _update_status_internal(self, message: str, verbose: bool):
        """Update status in main thread"""
        timestamp = datetime.now().strftime("[%H:%M:%S]")
        display_message = f"{timestamp} {message}"
        
        self.status_textbox.configure(state="normal")
        self.status_textbox.insert("end", "\n" + display_message)
        self.status_textbox.see("end")
        
        if not verbose:
            print(display_message)
        
        self.status_textbox.configure(state="disabled")

    # ==================================
    # Screen 2: Object Selection & Export
    # ==================================
    
    def _setup_export_frame(self):
        export_frame = self.export_frame
        export_frame.grid_rowconfigure(1, weight=1)  # Selection frame expands
        export_frame.grid_columnconfigure(0, weight=1)
        
        header_frame = ctk.CTkFrame(export_frame, fg_color="transparent")
        header_frame.grid(row=0, column=0, pady=(5, 5), sticky="ew")
        header_frame.columnconfigure(0, weight=1)
        ctk.CTkLabel(header_frame, text="Object Selection & Export", font=ctk.CTkFont(size=24, weight="bold")).grid(row=0, column=0, sticky="w")
        
        self.logout_button = ctk.CTkButton(header_frame, text="Logout", command=self.logout_action, width=100, fg_color="#CC3333")
        self.logout_button.grid(row=0, column=1, sticky="e", padx=10)

        selection_frame = ctk.CTkFrame(export_frame)
        selection_frame.grid(row=1, column=0, pady=5, sticky="nsew")
        selection_frame.grid_columnconfigure(0, weight=3)
        selection_frame.grid_columnconfigure(1, weight=1)
        selection_frame.grid_columnconfigure(2, weight=2)
        selection_frame.grid_rowconfigure(0, weight=1)
        
        # Available objects frame
        available_frame = ctk.CTkFrame(selection_frame)
        available_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        available_frame.grid_rowconfigure(2, weight=1)
        available_frame.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(available_frame, text="Available Objects", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, pady=3)
        
        self.search_entry = ctk.CTkEntry(available_frame, placeholder_text="Search...", height=30)
        self.search_entry.grid(row=1, column=0, padx=8, pady=3, sticky="ew")
        
        # Debounced search
        self._search_after_id = None
        self.search_entry.bind("<KeyRelease>", self._debounced_search)

        self.available_listbox = tk.Listbox(available_frame, selectmode="extended", height=12, exportselection=False,
                                            font=("Arial", 10), borderwidth=0, highlightthickness=0,
                                            selectbackground="#1F538D", fg="white", background="#242424")
        self.available_listbox.grid(row=2, column=0, padx=8, pady=(0, 8), sticky="nsew")

        # Action buttons frame
        action_frame = ctk.CTkFrame(selection_frame, fg_color="transparent")
        action_frame.grid(row=0, column=1, padx=3, pady=5, sticky="n")
        
        ctk.CTkLabel(action_frame, text="Actions", font=ctk.CTkFont(size=13, weight="bold")).pack(pady=3)
        
        ctk.CTkButton(action_frame, text=">> Add >>", command=self.add_selected_to_export, height=30, width=100).pack(pady=3, padx=3, fill="x")
        ctk.CTkButton(action_frame, text="<< Remove <<", command=self.remove_selected_from_export, height=30, width=100).pack(pady=3, padx=3, fill="x")

        ctk.CTkButton(action_frame, text="Select All", command=self.select_all_available, height=30, width=100).pack(pady=(15, 3), padx=3, fill="x")
        ctk.CTkButton(action_frame, text="Deselect All", command=self.deselect_all_available, height=30, width=100).pack(pady=3, padx=3, fill="x")

        # Selected objects frame
        selected_frame = ctk.CTkFrame(selection_frame)
        selected_frame.grid(row=0, column=2, padx=5, pady=5, sticky="nsew")
        selected_frame.grid_rowconfigure(1, weight=1)
        selected_frame.grid_columnconfigure(0, weight=1)
        
        header_container = ctk.CTkFrame(selected_frame, fg_color="transparent")
        header_container.grid(row=0, column=0, pady=3, sticky="ew")
        header_container.columnconfigure(0, weight=1)
        
        ctk.CTkLabel(header_container, text="Selected for Export", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, sticky="w", padx=8)
        
        self.selected_count_label = ctk.CTkLabel(header_container, text="(0)", font=ctk.CTkFont(size=13))
        self.selected_count_label.grid(row=0, column=1, sticky="e", padx=8)
        
        self.selected_listbox = tk.Listbox(selected_frame, selectmode="extended", height=12, exportselection=False,
                                           font=("Arial", 10), borderwidth=0, highlightthickness=0,
                                           selectbackground="#3366CC", fg="white", background="#242424")
        self.selected_listbox.grid(row=1, column=0, padx=8, pady=(0, 8), sticky="nsew")

        # Status textbox - FIXED HEIGHT
        status_label = ctk.CTkLabel(export_frame, text="Export Status:", font=ctk.CTkFont(size=14, weight="bold"), anchor="w")
        status_label.grid(row=2, column=0, padx=20, pady=(5, 0), sticky="w")
        
        self.status_textbox = ctk.CTkTextbox(export_frame, height=120)
        self.status_textbox.grid(row=3, column=0, padx=20, pady=(0, 5), sticky="ew")
        self.status_textbox.insert("end", "Status: Ready to select objects and export.")
        self.status_textbox.configure(state="disabled")

        # Export button frame with cancel button
        button_frame = ctk.CTkFrame(export_frame, fg_color="transparent")
        button_frame.grid(row=4, column=0, pady=(5, 10), sticky="ew", padx=20)
        button_frame.columnconfigure(0, weight=1)
        
        self.export_button = ctk.CTkButton(button_frame, text="Export Picklist Data", command=self.export_action, 
                                          height=45, fg_color="green", font=ctk.CTkFont(size=15, weight="bold"))
        self.export_button.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        
        self.cancel_button = ctk.CTkButton(button_frame, text="Cancel Export", command=self.cancel_export_action,
                                          height=45, fg_color="#CC3333", font=ctk.CTkFont(size=15, weight="bold"), width=150)
        self.cancel_button.grid(row=0, column=1, sticky="e")
        self.cancel_button.grid_remove()  # Hide initially
    
    def _debounced_search(self, event):
        """Debounce search to prevent excessive filtering"""
        if self._search_after_id:
            self.after_cancel(self._search_after_id)
        self._search_after_id = self.after(300, self.filter_available_objects)
    
    def update_status(self, message: str, verbose: bool = False):
        """Public method for status updates"""
        self._update_status_internal(message, verbose)

    # --- Object List Management Methods ---

    def populate_available_objects(self, objects: List[str]):
        """Populates the Left ListBox - optimized with batch updates"""
        self.available_listbox.delete(0, END)
        
        # Batch insert for better performance
        for obj in objects:
            self.available_listbox.insert(END, obj)
        
        # Color selected items in a separate pass
        for idx, obj in enumerate(objects):
            if obj in self.selected_objects:
                self.available_listbox.itemconfig(idx, {'fg': '#87CEEB'})

    def populate_selected_objects(self):
        """Populates the Right ListBox - optimized"""
        self.selected_listbox.delete(0, END)
        sorted_selection = sorted(list(self.selected_objects))
        
        for obj in sorted_selection:
            self.selected_listbox.insert(END, obj)
        
        # Update count label
        self.selected_count_label.configure(text=f"({len(self.selected_objects)})")

    def filter_available_objects(self):
        """Filters the Available ListBox - optimized"""
        search_term = self.search_entry.get().lower()
        
        if not search_term:
            # No filter - show all
            filtered_objects = self.all_org_objects
        else:
            # Filter with list comprehension (faster than loop)
            filtered_objects = [obj for obj in self.all_org_objects if search_term in obj.lower()]
        
        self.populate_available_objects(filtered_objects)
    
    def add_selected_to_export(self):
        """Adds selected objects - optimized"""
        selected_indices = self.available_listbox.curselection()
        
        if not selected_indices:
            messagebox.showwarning("Selection", "Please select one or more objects from the 'Available Objects' list to add.")
            return

        # Batch add to set
        objects_to_add = [self.available_listbox.get(i) for i in selected_indices]
        added_count = 0
        
        for obj_name in objects_to_add:
            if obj_name not in self.selected_objects:
                self.selected_objects.add(obj_name)
                added_count += 1
        
        if added_count > 0:
            self.populate_selected_objects()
            self.filter_available_objects()
            self.update_status(f"Added {added_count} object(s) to export list.")

    def remove_selected_from_export(self):
        """Removes selected objects - optimized"""
        selected_indices = self.selected_listbox.curselection()
        
        if not selected_indices:
            messagebox.showwarning("Selection", "Please select one or more objects from the 'Selected for Export' list to remove.")
            return

        # Batch remove from set
        objects_to_remove = [self.selected_listbox.get(i) for i in selected_indices]
        
        for obj_name in objects_to_remove:
            self.selected_objects.discard(obj_name)
        
        if objects_to_remove:
            self.populate_selected_objects()
            self.filter_available_objects()
            self.update_status(f"Removed {len(objects_to_remove)} object(s) from export list.")

    def select_all_available(self):
        """Selects all objects currently visible"""
        self.available_listbox.select_set(0, END)
    
    def deselect_all_available(self):
        """Deselects all objects"""
        self.available_listbox.select_clear(0, END)

    # --- Action Methods ---

    def logout_action(self):
        """Logout with confirmation - PRESERVES INPUT FIELDS"""
        if self.is_exporting:
            messagebox.showwarning("Export In Progress", "Cannot logout while export is running. Please wait or cancel the export.")
            return
            
        confirm = messagebox.askyesno("Logout", "Are you sure you want to log out?")
        if confirm:
            self.sf_exporter = None
            self.selected_objects.clear()
            self.all_org_objects.clear()
            
            # Reset login button
            self.login_button.configure(state="normal", text="Login to Salesforce")
            
            # DO NOT clear credentials - keep them for re-login
            # self.username_entry.delete(0, END)  # REMOVED
            # self.password_entry.delete(0, END)  # REMOVED
            # self.token_entry.delete(0, END)     # REMOVED
            
            self.update_status("Logged out successfully. Please log in again.")
            
            # Switch back to Login Frame
            self.export_frame.grid_forget()
            self.login_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

    def export_action(self):
        """Export with threading to prevent UI freeze"""
        if not self.sf_exporter:
            messagebox.showerror("Error", "Not logged in. Please log in first.")
            return

        if self.is_exporting:
            messagebox.showwarning("Export In Progress", "An export is already running. Please wait.")
            return

        selected_objects_list = sorted(list(self.selected_objects))

        if not selected_objects_list:
            messagebox.showwarning("Warning", "The 'Selected for Export' list is empty. Please add objects.")
            return

        default_filename = f'Picklist_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not output_file_path:
            return

        # Update UI state
        self.is_exporting = True
        self.export_button.configure(state="disabled", text="Exporting...")
        self.cancel_button.grid()  # Show cancel button
        self.logout_button.configure(state="disabled")
        
        self.update_status(f"Starting export for {len(selected_objects_list)} objects to {output_file_path}...")
        
        # Start export in background thread
        def export_worker():
            start_time = time.time()
            output_path = None
            stats = None
            
            try:
                output_path, stats = self.sf_exporter.export_picklists(
                    selected_objects_list, 
                    output_file_path
                )
                
                end_time = time.time()
                runtime_seconds = end_time - start_time
                runtime_formatted = format_runtime(runtime_seconds)
                
                # Queue completion message
                self.message_queue.put(("export_complete", (output_path, stats, runtime_formatted)))
                
            except Exception as e:
                # Queue error message
                self.message_queue.put(("export_error", str(e)))
        
        self.export_thread = threading.Thread(target=export_worker, daemon=True)
        self.export_thread.start()

    def cancel_export_action(self):
        """Cancel ongoing export"""
        if not self.is_exporting:
            return
        
        confirm = messagebox.askyesno("Cancel Export", "Are you sure you want to cancel the export?")
        if confirm and self.sf_exporter:
            self.sf_exporter.cancel_export()
            self.update_status("⚠️ Cancelling export... Please wait.")

    def _handle_export_complete(self, output_path, stats, runtime):
        """Handle export completion in main thread"""
        self.is_exporting = False
        self.export_button.configure(state="normal", text="Export Picklist Data")
        self.cancel_button.grid_remove()  # Hide cancel button
        self.logout_button.configure(state="normal")
        
        if stats.get('cancelled', False):
            self.update_status(f"⚠️ Export Cancelled. Partial Runtime: {runtime}")
            messagebox.showwarning("Export Cancelled", "The export was cancelled by user request.")
        else:
            self.update_status(f"✅ Export Complete! Total Runtime: {runtime}")
            messagebox.showinfo("Export Done", f"Picklist data successfully exported to:\n{output_path}")
        
        # Print statistics to console
        print_statistics(stats, runtime, output_path)

    def _handle_export_error(self, error_msg):
        """Handle export error in main thread"""
        self.is_exporting = False
        self.export_button.configure(state="normal", text="Export Picklist Data")
        self.cancel_button.grid_remove()
        self.logout_button.configure(state="normal")
        
        self.update_status(f"❌ FATAL EXPORT ERROR: {error_msg}")
        messagebox.showerror("Export Error", f"A fatal error occurred during export:\n{error_msg}")

    def on_closing(self):
        """Handle window close event"""
        if self.is_exporting:
            confirm = messagebox.askyesno(
                "Export In Progress", 
                "An export is currently running. Are you sure you want to close?\nThis will cancel the export."
            )
            if not confirm:
                return
            
            if self.sf_exporter:
                self.sf_exporter.cancel_export()
        
        self.destroy()


# ===========================================
# MAIN EXECUTION
# ===========================================

if __name__ == "__main__":
    try:
        app = PicklistExportGUI()
        app.protocol("WM_DELETE_WINDOW", app.on_closing)  # Handle window close
        app.mainloop()
    except Exception as e:
        print(f"\n❌ GUI Application Failed: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)