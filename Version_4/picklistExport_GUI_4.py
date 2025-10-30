'''
Prerequisites (Installation)
   pip install simple-salesforce openpyxl requests customtkinter
   
Single Command Installation:
   pip install simple-salesforce openpyxl requests customtkinter

'''
import os
import sys
import time
import requests
import tkinter as tk 
import threading
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple, Set

# Third-party libraries
from simple_salesforce import Salesforce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# GUI Library
import customtkinter as ctk
from tkinter import messagebox, filedialog, END, ttk

# Set appearance mode and default color theme
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

# ===========================================
# CONFIGURATION / GLOBAL CONSTANTS
# ===========================================

API_VERSION = '65.0'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ===========================================
# HELPER CLASSES 
# ===========================================

class FieldInfo:
    """Represents picklist field metadata"""
    def __init__(self, api_name: str, label: str):
        self.api_name = api_name
        self.label = label


class PicklistValueDetail:
    """Represents a single picklist value"""
    def __init__(self, label: str, value: str, is_active: bool = True):
        self.label = label
        self.value = value
        self.is_active = is_active


class ProcessingResult:
    """Stores processing results for an object"""
    def __init__(self):
        self.values_processed = 0
        self.inactive_values = 0
        self.rows: List[List[str]] = []
        self.picklist_fields_count = 0
        self.object_exists = True
        self.error_message = None


# ===========================================
# MAIN EXPORT CLASS 
# ===========================================

class PicklistExporter:
    """Main exporter class with enhanced statistics"""
    
    def __init__(self, username: str, password: str, security_token: str, domain: str = 'login', status_callback=None):
        """Initialize Salesforce connection"""
        self.status_callback = status_callback
        self.all_org_objects: List[str] = []
        
        if self.status_callback:
            self.status_callback("Initializing Salesforce Connection...")
            
        try:
            self.sf = Salesforce(
                username=username,
                password=password,
                security_token=security_token,
                domain=domain
            )
            self.base_url = f"https://{self.sf.sf_instance}"
            self.session_id = self.sf.session_id
            self.headers = {
                'Authorization': f'Bearer {self.session_id}',
                'Content-Type': 'application/json'
            }
            if self.status_callback:
                self.status_callback(f"✅ Connected to: {self.base_url}")
            
            self._fetch_all_org_objects()

        except Exception as e:
            if self.status_callback:
                self.status_callback(f"❌ Connection failed: {str(e)}")
            raise

    def _fetch_all_org_objects(self):
        """Fetches all SObjects (Standard and Custom) from the org"""
        self._log_status("Fetching all available SObjects from the organization...")
        try:
            response = self.sf.describe()
            self.all_org_objects = sorted([
                obj['name'] for obj in response['sobjects'] 
                if obj.get('queryable', False) and not obj.get('deprecatedAndHidden', False)
            ])
            self._log_status(f"✅ Found {len(self.all_org_objects)} queryable objects.")
        except Exception as e:
            self._log_status(f"❌ Failed to fetch all SObjects: {str(e)}")
            self.all_org_objects = []
            
    def get_all_objects(self) -> List[str]:
        """Accessor for the fetched object list"""
        return self.all_org_objects

    def _log_status(self, message: str):
        """Internal helper to send log messages back to the GUI"""
        if self.status_callback:
            self.status_callback(message, verbose=True) 

    def export_picklists(self, object_names: List[str], output_path: str, progress_callback=None) -> Tuple[str, Dict]:
        self._log_status("=== Starting Picklist Export ===")
        self._log_status(f"Total objects to process: {len(object_names)}")
        
        stats = {
            'total_objects': len(object_names), 'successful_objects': 0, 'failed_objects': 0, 'objects_not_found': 0,
            'objects_with_zero_picklists': 0, 'objects_with_picklists': 0, 'total_picklist_fields': 0, 'total_values': 0,
            'total_active_values': 0, 'total_inactive_values': 0, 'failed_object_details': [],
            'objects_without_picklists': [], 'objects_not_found_list': []
        }
        
        all_rows = [['Object', 'Field Label', 'Field API', 'Picklist Value Label', 'Picklist Value API', 'Status']]
        
        for i, obj_name in enumerate(object_names, 1):
            if progress_callback:
                progress_callback(i, len(object_names))
            
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({'name': obj_name, 'reason': 'Object does not exist in org'})
                    self._log_status(f"  ⚠️  Object not found in org")
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  ℹ️  No picklist fields found")
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    all_rows.extend(result.rows)
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    self._log_status(f"  ✅ Fields: {result.picklist_fields_count}, Active: {result.values_processed - result.inactive_values}, Inactive: {result.inactive_values}")
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({'name': obj_name, 'reason': error_msg})
            self._log_status("")
        
        self._log_status("=== Creating Excel File ===")
        final_output_path = self._create_excel_file(all_rows, output_path)
        return final_output_path, stats
    
    def _process_object(self, obj_name: str) -> ProcessingResult:
        result = ProcessingResult()
        try:
            getattr(self.sf, obj_name).describe()
        except Exception as e:
            if 'NOT_FOUND' in str(e) or 'INVALID_TYPE' in str(e):
                result.object_exists = False
                return result
            raise
        
        picklist_fields = self._get_picklist_fields(obj_name)
        result.picklist_fields_count = len(picklist_fields)
        if not picklist_fields: return result
        self._log_status(f"  Found {len(picklist_fields)} picklist fields")
        entity_def_id = self._resolve_entity_definition_id(obj_name)
        if entity_def_id: self._log_status(f"  EntityDefinition.Id: {entity_def_id}")
        
        for field_api, field_info in picklist_fields.items():
            values = self._query_picklist_values_with_fallback(obj_name, entity_def_id, field_api)
            if not values: continue
            self._log_status(f"    Field: {field_api} - {len(values)} values")
            for value in values:
                is_active = value.is_active if value.is_active is not None else True
                status = 'Active' if is_active else 'Inactive'
                if not is_active: result.inactive_values += 1
                row = [obj_name, field_info.label, field_api, value.label, value.value, status]
                result.rows.append(row)
                result.values_processed += 1
        return result
    
    def _get_picklist_fields(self, object_name: str) -> Dict[str, FieldInfo]:
        fields_dict = {}
        try:
            obj_describe = getattr(self.sf, object_name).describe()
            for field in obj_describe['fields']:
                if field['type'] in ['picklist', 'multipicklist']:
                    fields_dict[field['name']] = FieldInfo(api_name=field['name'], label=field['label'])
        except Exception as e:
            self._log_status(f"  ERROR in _get_picklist_fields: {str(e)}")
        return fields_dict
    
    def _resolve_entity_definition_id(self, object_name: str) -> Optional[str]:
        try:
            query = f"SELECT Id FROM EntityDefinition WHERE QualifiedApiName = '{object_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records: return records[0]['Id']
        except Exception as e:
            self._log_status(f"  ERROR resolveEntityDefinitionId: {str(e)}")
        return None
    
    def _query_picklist_values_with_fallback(self, object_name: str, entity_def_id: Optional[str], field_name: str) -> List[PicklistValueDetail]:
        values = self._query_field_definition_tooling(object_name, field_name)
        if values: return values
        if entity_def_id:
            values = self._query_custom_field_tooling(entity_def_id, field_name)
            if values: return values
        values = self._query_custom_field_tooling_table_enum(object_name, field_name)
        if values: return values
        values = self._query_rest_describe_for_picklist(object_name, field_name)
        if values: return values
        return []
    
    def _query_field_definition_tooling(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        try:
            query = f"SELECT Metadata FROM FieldDefinition WHERE EntityDefinition.QualifiedApiName = '{object_name}' AND QualifiedApiName = '{field_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records: return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryFieldDefinitionTooling: {str(e)}")
        return []
    
    def _query_custom_field_tooling(self, entity_def_id: str, field_name: str) -> List[PicklistValueDetail]:
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{entity_def_id}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records: return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryCustomFieldTooling: {str(e)}")
        return []
    
    def _query_custom_field_tooling_table_enum(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{object_name}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records: return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryCustomFieldToolingTableEnum: {str(e)}")
        return []
    
    def _query_rest_describe_for_picklist(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        try:
            url = f"{self.base_url}/services/data/v{API_VERSION}/sobjects/{object_name}/describe"
            response = requests.get(url, headers=self.headers, timeout=60)
            if response.status_code == 200:
                for field in response.json().get('fields', []):
                    if field['name'].lower() == field_name.lower():
                        return [PicklistValueDetail(label=pv.get('label', ''), value=pv.get('value', ''), is_active=pv.get('active', True))
                                for pv in field.get('picklistValues', [])]
        except Exception as e:
            self._log_status(f"      ERROR queryRestDescribeForPicklist: {str(e)}")
        return []
    
    def _parse_value_set(self, metadata: dict) -> List[PicklistValueDetail]:
        results = []
        try:
            value_set = metadata.get('valueSet', {})
            if not value_set: return results
            values = value_set.get('valueSetDefinition', {}).get('value', []) or value_set.get('value', [])
            for v in values:
                is_active = bool(v.get('isActive', True))
                results.append(PicklistValueDetail(label=v.get('label', ''), value=v.get('valueName') or v.get('value', ''), is_active=is_active))
        except Exception as e:
            self._log_status(f"      ERROR parseValueSet: {str(e)}")
        return results
    
    def _create_excel_file(self, rows: List[List[str]], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Picklist Export"
        for row in rows: ws.append(row)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        ws.freeze_panes = "A2"
        wb.save(output_path)
        self._log_status(f"✅ Excel file created: {output_path}")
        self._log_status(f"✅ Total data rows: {len(rows) - 1}")
        return output_path


# ===========================================
# GUI HELPER FUNCTIONS
# ===========================================

def format_runtime(seconds: float) -> str:
    """Format runtime in HH:MM:SS format"""
    td = timedelta(seconds=int(seconds))
    hours, remainder = divmod(td.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def print_statistics(stats: Dict, runtime_formatted: str, output_file: str):
    """Prints comprehensive statistics to the console"""
    print("\n" + "=" * 70)
    print("✅ EXPORT COMPLETED SUCCESSFULLY! (Statistics Detail)")
    print("=" * 70)
    print(f"Total Runtime: {runtime_formatted}")
    print(f"Total Objects in List:          {stats['total_objects']}")
    print(f"✅ Successfully Processed:       {stats['successful_objects']}")
    print(f"❌ Failed to Process:            {stats['failed_objects']}")
    print(f"⚠️  Objects Not Found in Org:    {stats['objects_not_found']}")
    print(f"Total Picklist Fields:          {stats['total_picklist_fields']}")
    print(f"Total Picklist Values:          {stats['total_values']}")
    print(f"✅ Active Values:                {stats['total_active_values']}")
    print(f"❌ Inactive Values:              {stats['total_inactive_values']}")
    print(f"Output File: {output_file}")
    if stats['failed_objects'] > 0:
        print("\n❌ FAILED OBJECTS (REASONS):")
        for detail in stats['failed_object_details']:
            print(f"   • {detail['name']}: {detail['reason']}")
    print("=" * 70)


# ===========================================
# GUI IMPLEMENTATION
# ===========================================

class PicklistExportGUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("Salesforce Picklist Exporter")
        self.geometry("1200x720")
        self.resizable(False, False)
        
        self.sf_exporter: Optional[PicklistExporter] = None
        self.all_org_objects: List[str] = []
        self.selected_objects: Set[str] = set()
        self.current_filter = "all"  # all, standard, custom
        
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        self.login_frame = ctk.CTkFrame(self)
        self.export_frame = ctk.CTkFrame(self)
        
        self.login_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        self.export_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        
        self._setup_login_frame()
        self._setup_export_frame()
        
        self.export_frame.grid_forget()

    # ==================================
    # Screen 1: Login & Authentication
    # ==================================
    
    def _setup_login_frame(self):
        login_frame = self.login_frame
        login_frame.columnconfigure(1, weight=1)
        
        ctk.CTkLabel(login_frame, text="Salesforce Login", font=ctk.CTkFont(size=30, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(50, 40))

        def create_input_row(parent, row, label_text, password_mode=False):
            ctk.CTkLabel(parent, text=label_text, anchor="w", font=ctk.CTkFont(size=14)).grid(row=row, column=0, padx=10, pady=15, sticky="w")
            entry = ctk.CTkEntry(parent, width=350, show="*" if password_mode else "")
            entry.grid(row=row, column=1, padx=10, pady=15, sticky="ew")
            return entry

        self.username_entry = create_input_row(login_frame, 1, "Username:")
        self.password_entry = create_input_row(login_frame, 2, "Password:", password_mode=True)
        self.token_entry = create_input_row(login_frame, 3, "Security Token:", password_mode=True)

        ctk.CTkLabel(login_frame, text="Org Type:", anchor="w", font=ctk.CTkFont(size=14)).grid(row=4, column=0, padx=10, pady=15, sticky="w")
        self.org_type_var = ctk.StringVar(value="Production")
        radio_prod = ctk.CTkRadioButton(login_frame, text="Production", variable=self.org_type_var, value="Production")
        radio_test = ctk.CTkRadioButton(login_frame, text="Sandbox/Test", variable=self.org_type_var, value="Sandbox")
        
        radio_prod.grid(row=4, column=1, padx=(10, 5), pady=15, sticky="w")
        radio_test.grid(row=4, column=1, padx=(140, 10), pady=15, sticky="w")
        
        self.login_button = ctk.CTkButton(login_frame, text="Login to Salesforce", command=self.login_action, height=50, font=ctk.CTkFont(size=16, weight="bold"))
        self.login_button.grid(row=5, column=0, columnspan=2, pady=50, sticky="ew", padx=10)

    def login_action(self):
        self.login_button.configure(state="disabled", text="Connecting...")
        
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        token = self.token_entry.get().strip()
        domain = 'test' if self.org_type_var.get() == 'Sandbox' else 'login'
        
        if not all([username, password, token]):
            messagebox.showerror("Input Error", "All fields (Username, Password, Security Token) are required.")
            self.login_button.configure(state="normal", text="Login to Salesforce")
            return

        try:
            self.sf_exporter = PicklistExporter(
                username=username, 
                password=password, 
                security_token=token, 
                domain=domain,
                status_callback=self.update_status
            )
            
            messagebox.showinfo("Success", "Successfully connected to Salesforce!")
            
            self.all_org_objects = self.sf_exporter.get_all_objects()
            
            self.login_frame.grid_forget()
            self.export_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
            self.populate_available_objects(self.all_org_objects)
            self.populate_selected_objects()
            self.update_object_counts()
            
        except Exception as e:
            messagebox.showerror("Login Failed", f"Connection Error: {str(e)}")
            self.sf_exporter = None
            self.login_button.configure(state="normal", text="Login to Salesforce")

    # ==================================
    # Screen 2: Object Selection & Export
    # ==================================
    
    def _setup_export_frame(self):
        export_frame = self.export_frame
        export_frame.grid_rowconfigure(2, weight=1)
        export_frame.grid_columnconfigure(0, weight=1)
        
        # Header with title, theme toggle, and logout
        header_frame = ctk.CTkFrame(export_frame, fg_color="transparent")
        header_frame.grid(row=0, column=0, pady=(10, 5), sticky="ew", padx=10)
        header_frame.columnconfigure(1, weight=1)
        
        ctk.CTkLabel(header_frame, text="Salesforce Picklist Exporter", font=ctk.CTkFont(size=28, weight="bold")).grid(row=0, column=0, sticky="w")
        
        # Theme toggle button
        self.theme_toggle = ctk.CTkButton(
            header_frame, 
            text="🌙", 
            command=self.toggle_theme, 
            width=40,
            height=40,
            font=ctk.CTkFont(size=20)
        )
        self.theme_toggle.grid(row=0, column=1, sticky="e", padx=(0, 10))
        
        self.logout_button = ctk.CTkButton(header_frame, text="Logout", command=self.logout_action, width=100, fg_color="#CC3333", height=40)
        self.logout_button.grid(row=0, column=2, sticky="e")

        # Object selection panel
        selection_frame = ctk.CTkFrame(export_frame)
        selection_frame.grid(row=1, column=0, pady=10, sticky="nsew", padx=10)
        selection_frame.grid_columnconfigure(0, weight=3)
        selection_frame.grid_columnconfigure(1, weight=1)
        selection_frame.grid_columnconfigure(2, weight=2)
        selection_frame.grid_rowconfigure(0, weight=1)
        
        # LEFT: Available Objects
        available_frame = ctk.CTkFrame(selection_frame)
        available_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        available_frame.grid_rowconfigure(3, weight=1)
        available_frame.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(available_frame, text="Available Objects", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, pady=(5, 2))
        
        self.available_count_label = ctk.CTkLabel(available_frame, text="(0 objects)", font=ctk.CTkFont(size=11))
        self.available_count_label.grid(row=1, column=0, pady=(0, 5))
        
        # Filter buttons
        filter_frame = ctk.CTkFrame(available_frame, fg_color="transparent")
        filter_frame.grid(row=2, column=0, pady=5, sticky="ew", padx=5)
        
        self.filter_all_btn = ctk.CTkButton(filter_frame, text="All", command=lambda: self.apply_filter("all"), width=60, height=25)
        self.filter_all_btn.grid(row=0, column=0, padx=2)
        
        self.filter_standard_btn = ctk.CTkButton(filter_frame, text="Standard", command=lambda: self.apply_filter("standard"), width=70, height=25, fg_color="gray")
        self.filter_standard_btn.grid(row=0, column=1, padx=2)
        
        self.filter_custom_btn = ctk.CTkButton(filter_frame, text="Custom", command=lambda: self.apply_filter("custom"), width=70, height=25, fg_color="gray")
        self.filter_custom_btn.grid(row=0, column=2, padx=2)
        
        self.search_entry = ctk.CTkEntry(available_frame, placeholder_text="Search Object API Name...", height=30)
        self.search_entry.grid(row=3, column=0, padx=5, pady=5, sticky="ew")
        self.search_entry.bind("<KeyRelease>", self.filter_available_objects)

        self.available_listbox = tk.Listbox(
            available_frame, selectmode="extended", height=10, exportselection=False,
            font=("Arial", 11), borderwidth=0, highlightthickness=0,
            selectbackground="#1F538D", fg="white", background="#2B2B2B"
        )
        self.available_listbox.grid(row=4, column=0, padx=5, pady=(0, 5), sticky="nsew")

        # MIDDLE: Action buttons
        action_frame = ctk.CTkFrame(selection_frame, fg_color="transparent")
        action_frame.grid(row=0, column=1, padx=3, pady=5)
        
        ctk.CTkLabel(action_frame, text="Actions", font=ctk.CTkFont(size=13, weight="bold")).pack(pady=5)
        
        ctk.CTkButton(action_frame, text="Add >>", command=self.add_selected_to_export, height=30, width=90).pack(pady=3, padx=3)
        ctk.CTkButton(action_frame, text="<< Remove", command=self.remove_selected_from_export, height=30, width=90).pack(pady=3, padx=3)
        ctk.CTkButton(action_frame, text="Select All", command=self.select_all_available, height=30, width=90).pack(pady=(15, 3), padx=3)
        ctk.CTkButton(action_frame, text="Deselect All", command=self.deselect_all_available, height=30, width=90).pack(pady=3, padx=3)

        # RIGHT: Selected Objects
        selected_frame = ctk.CTkFrame(selection_frame)
        selected_frame.grid(row=0, column=2, padx=5, pady=5, sticky="nsew")
        selected_frame.grid_rowconfigure(2, weight=1)
        selected_frame.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(selected_frame, text="Selected for Export", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, pady=(5, 2))
        
        self.selected_count_label = ctk.CTkLabel(selected_frame, text="(0 selected)", font=ctk.CTkFont(size=11))
        self.selected_count_label.grid(row=1, column=0, pady=(0, 5))
        
        self.selected_listbox = tk.Listbox(
            selected_frame, selectmode="extended", height=10, exportselection=False,
            font=("Arial", 11), borderwidth=0, highlightthickness=0,
            selectbackground="#3366CC", fg="white", background="#2B2B2B"
        )
        self.selected_listbox.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="nsew")

        # Export button
        self.export_button = ctk.CTkButton(
            export_frame, 
            text="Export Picklist Data", 
            command=self.export_action, 
            height=45, 
            fg_color="#28a745",
            hover_color="#218838",
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.export_button.grid(row=2, column=0, pady=(5, 5), sticky="ew", padx=10)

        # Status bar (green)
        self.status_bar = ctk.CTkLabel(
            export_frame, 
            text="Status: Ready", 
            font=ctk.CTkFont(size=12),
            fg_color="#28a745",
            height=30,
            anchor="w",
            padx=10
        )
        self.status_bar.grid(row=3, column=0, sticky="ew", padx=10, pady=(5, 0))

        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(export_frame, height=15)
        self.progress_bar.grid(row=4, column=0, sticky="ew", padx=10, pady=(5, 0))
        self.progress_bar.set(0)

        # Terminal/Log area (40% of window height = ~288px)
        self.status_textbox = ctk.CTkTextbox(export_frame, height=280, font=("Consolas", 10))
        self.status_textbox.grid(row=5, column=0, padx=10, pady=(5, 10), sticky="nsew")
        self.status_textbox.insert("end", "═══════════════════════════════════════════════════════════════\n")
        self.status_textbox.insert("end", "  Salesforce Picklist Exporter - Ready\n")
        self.status_textbox.insert("end", "═══════════════════════════════════════════════════════════════\n")
        self.status_textbox.insert("end", "\n✓ Connected successfully. Select objects to export.\n")
        self.status_textbox.configure(state="disabled")
    
    def toggle_theme(self):
        """Toggle between dark and light mode"""
        current_mode = ctk.get_appearance_mode()
        if current_mode == "Dark":
            ctk.set_appearance_mode("Light")
            self.theme_toggle.configure(text="☀️")
            # Update listbox colors for light mode
            self.available_listbox.configure(fg="black", background="#F0F0F0")
            self.selected_listbox.configure(fg="black", background="#F0F0F0")
        else:
            ctk.set_appearance_mode("Dark")
            self.theme_toggle.configure(text="🌙")
            # Update listbox colors for dark mode
            self.available_listbox.configure(fg="white", background="#2B2B2B")
            self.selected_listbox.configure(fg="white", background="#2B2B2B")
    
    def update_status(self, message: str, verbose: bool = False):
        """Updates the GUI status text box with new messages"""
        timestamp = datetime.now().strftime("[%H:%M:%S]")
        display_message = f"{timestamp} {message}"
        
        self.status_textbox.configure(state="normal")
        self.status_textbox.insert("end", "\n" + display_message)
        self.status_textbox.see("end")
        
        if not verbose:
            print(display_message) 

        self.status_textbox.configure(state="disabled")
        self.update_idletasks()
    
    def update_status_bar(self, message: str, color: str = "#28a745"):
        """Updates the status bar with a message and color"""
        self.status_bar.configure(text=f"Status: {message}", fg_color=color)
        self.update_idletasks()
    
    def update_progress(self, current: int, total: int):
        """Updates the progress bar"""
        if total > 0:
            progress = current / total
            self.progress_bar.set(progress)
            self.update_status_bar(f"Processing {current}/{total} objects...", "#FFA500")
        self.update_idletasks()

    # --- Object List Management Methods ---

    def apply_filter(self, filter_type: str):
        """Apply object filter (all/standard/custom)"""
        self.current_filter = filter_type
        
        # Update button colors
        self.filter_all_btn.configure(fg_color="gray" if filter_type != "all" else None)
        self.filter_standard_btn.configure(fg_color="gray" if filter_type != "standard" else None)
        self.filter_custom_btn.configure(fg_color="gray" if filter_type != "custom" else None)
        
        self.filter_available_objects(None)
    
    def get_filtered_objects(self) -> List[str]:
        """Get objects based on current filter"""
        if self.current_filter == "standard":
            return [obj for obj in self.all_org_objects if not obj.endswith('__c')]
        elif self.current_filter == "custom":
            return [obj for obj in self.all_org_objects if obj.endswith('__c')]
        else:
            return self.all_org_objects

    def populate_available_objects(self, objects: List[str]):
        """Populates the Left ListBox based on the current search filter."""
        self.available_listbox.delete(0, END)
        for obj in objects:
            self.available_listbox.insert(END, obj)
            if obj in self.selected_objects:
                idx = self.available_listbox.get(0, END).index(obj)
                self.available_listbox.itemconfig(idx, {'fg': '#87CEEB'})

    def populate_selected_objects(self):
        """Populates the Right ListBox from the internal selected_objects set (always sorted)."""
        self.selected_listbox.delete(0, END)
        for obj in sorted(list(self.selected_objects)):
            self.selected_listbox.insert(END, obj)

    def filter_available_objects(self, event):
        """Filters the Available ListBox based on the search entry content and current filter."""
        search_term = self.search_entry.get().lower()
        base_objects = self.get_filtered_objects()
        
        filtered_objects = [
            obj for obj in base_objects
            if search_term in obj.lower()
        ]
        self.populate_available_objects(filtered_objects)
        self.update_object_counts()
    
    def update_object_counts(self):
        """Update the count labels"""
        available_count = self.available_listbox.size()
        selected_count = len(self.selected_objects)
        
        self.available_count_label.configure(text=f"({available_count} objects)")
        self.selected_count_label.configure(text=f"({selected_count} selected)")
    
    def add_selected_to_export(self):
        """Adds selected objects from the Available List to the Export Set."""
        selected_indices = self.available_listbox.curselection()
        
        if not selected_indices:
            messagebox.showwarning("Selection", "Please select one or more objects from the 'Available Objects' list to add.")
            return

        added_count = 0
        for i in selected_indices:
            obj_name = self.available_listbox.get(i)
            if obj_name not in self.selected_objects:
                self.selected_objects.add(obj_name)
                added_count += 1
        
        if added_count > 0:
            self.populate_selected_objects()
            self.filter_available_objects(None)
            self.update_status(f"✓ Added {added_count} object(s) to export list.")
            self.update_object_counts()

    def remove_selected_from_export(self):
        """Removes selected objects from the Selected List and the Export Set."""
        selected_indices = self.selected_listbox.curselection()
        
        if not selected_indices:
            messagebox.showwarning("Selection", "Please select one or more objects from the 'Selected for Export' list to remove.")
            return

        removed_objects = []
        for i in reversed(selected_indices):
            obj_name = self.selected_listbox.get(i)
            removed_objects.append(obj_name)
        
        for obj_name in removed_objects:
            self.selected_objects.discard(obj_name)
        
        if removed_objects:
            self.populate_selected_objects()
            self.filter_available_objects(None)
            self.update_status(f"✓ Removed {len(removed_objects)} object(s) from export list.")
            self.update_object_counts()

    def select_all_available(self):
        """Selects all objects currently visible in the Available ListBox."""
        self.available_listbox.select_set(0, END)
    
    def deselect_all_available(self):
        """Deselects all objects currently visible in the Available ListBox."""
        self.available_listbox.select_clear(0, END)

    # --- Action Methods ---

    def logout_action(self):
        """Clears connection, resets state, and returns to the login screen."""
        confirm = messagebox.askyesno("Logout", "Are you sure you want to log out?")
        if confirm:
            self.sf_exporter = None
            self.selected_objects.clear()
            self.all_org_objects.clear()
            
            self.login_button.configure(state="normal", text="Login to Salesforce") 
            self.update_status("✓ Logged out successfully. Please log in again.")
            
            self.export_frame.grid_forget()
            self.login_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

    def export_action(self):
        """Export action with threading to prevent UI freeze"""
        if not self.sf_exporter:
            messagebox.showerror("Error", "Not logged in. Please log in first.")
            return

        selected_objects_list = sorted(list(self.selected_objects))

        if not selected_objects_list:
            messagebox.showwarning("Warning", "The 'Selected for Export' list is empty. Please add objects.")
            return

        default_filename = f'Picklist_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not output_file_path:
            return

        # Disable all interactive elements
        self.disable_ui()
        self.export_button.configure(text="Exporting... Please Wait")
        self.update_status_bar("Export in progress...", "#FFA500")
        self.progress_bar.set(0)
        
        # Run export in background thread
        export_thread = threading.Thread(
            target=self._run_export,
            args=(selected_objects_list, output_file_path),
            daemon=True
        )
        export_thread.start()

    def _run_export(self, selected_objects_list: List[str], output_file_path: str):
        """Background thread for export operation"""
        start_time = time.time()
        output_path = None
        stats = None
        
        try:
            output_path, stats = self.sf_exporter.export_picklists(
                selected_objects_list, 
                output_file_path,
                progress_callback=self.update_progress
            )
            
            end_time = time.time()
            runtime_seconds = end_time - start_time
            runtime_formatted = format_runtime(runtime_seconds)
            
            self.after(0, self._export_complete_success, output_path, stats, runtime_formatted)
            
        except Exception as e:
            self.after(0, self._export_complete_error, str(e))

    def _export_complete_success(self, output_path: str, stats: Dict, runtime_formatted: str):
        """Called when export completes successfully"""
        self.update_status(f"\n{'='*60}")
        self.update_status(f"✅ EXPORT COMPLETED SUCCESSFULLY!")
        self.update_status(f"{'='*60}")
        self.update_status(f"Total Runtime: {runtime_formatted}")
        self.update_status(f"Total Objects: {stats['total_objects']}")
        self.update_status(f"✅ Successful: {stats['successful_objects']}")
        self.update_status(f"❌ Failed: {stats['failed_objects']}")
        self.update_status(f"Total Picklist Fields: {stats['total_picklist_fields']}")
        self.update_status(f"Total Values: {stats['total_values']} (Active: {stats['total_active_values']}, Inactive: {stats['total_inactive_values']})")
        self.update_status(f"Output File: {output_path}")
        self.update_status(f"{'='*60}\n")
        
        self.update_status_bar("Export completed successfully!", "#28a745")
        self.progress_bar.set(1.0)
        
        messagebox.showinfo("Export Complete", f"Picklist data successfully exported!\n\nFile: {output_path}\nRuntime: {runtime_formatted}")
        
        print_statistics(stats, runtime_formatted, output_path)
        
        self.enable_ui()
        self.export_button.configure(text="Export Picklist Data")

    def _export_complete_error(self, error_message: str):
        """Called when export fails"""
        self.update_status(f"\n❌ FATAL EXPORT ERROR: {error_message}\n")
        self.update_status_bar("Export failed!", "#CC3333")
        
        messagebox.showerror("Export Error", f"A fatal error occurred during export:\n\n{error_message}")
        
        self.enable_ui()
        self.export_button.configure(text="Export Picklist Data")

    def disable_ui(self):
        """Disable all interactive UI elements during export"""
        self.available_listbox.configure(state="disabled")
        self.selected_listbox.configure(state="disabled")
        self.search_entry.configure(state="disabled")
        self.export_button.configure(state="disabled")
        self.logout_button.configure(state="disabled")
        self.theme_toggle.configure(state="disabled")
        self.filter_all_btn.configure(state="disabled")
        self.filter_standard_btn.configure(state="disabled")
        self.filter_custom_btn.configure(state="disabled")
        self.configure(cursor="watch")

    def enable_ui(self):
        """Re-enable all interactive UI elements after export"""
        self.available_listbox.configure(state="normal")
        self.selected_listbox.configure(state="normal")
        self.search_entry.configure(state="normal")
        self.export_button.configure(state="normal")
        self.logout_button.configure(state="normal")
        self.theme_toggle.configure(state="normal")
        self.filter_all_btn.configure(state="normal")
        self.filter_standard_btn.configure(state="normal")
        self.filter_custom_btn.configure(state="normal")
        self.configure(cursor="")

# ===========================================
# MAIN EXECUTION
# ===========================================

if __name__ == "__main__":
    try:
        app = PicklistExportGUI()
        app.mainloop()
    except Exception as e:
        print(f"\n❌ GUI Application Failed: {str(e)}")
        sys.exit(1)