import os
import sys
import time
import requests
import tkinter as tk # Explicitly import tk for Listbox
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple

# Third-party libraries
from simple_salesforce import Salesforce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# GUI Library
import customtkinter as ctk
from tkinter import messagebox, filedialog, END

# Set appearance mode and default color theme
ctk.set_appearance_mode("System")  # Modes: "System" (default), "Dark", "Light"
ctk.set_default_color_theme("blue") # Themes: "blue" (default), "green", "dark-blue"

# ===========================================
# CONFIGURATION - Used as initial list for GUI
# ===========================================

SF_USERNAME = ''
SF_PASSWORD = ''
SF_SECURITY_TOKEN = ''
SF_DOMAIN = 'login' # default to production

# Objects to export - This list populates the GUI ListBox
OBJECTS_TO_EXPORT = [
    'Account', 'Contact', 'Opportunity', 'Lead', 'Case', 'User', 
    'Product2', 'Asset', 'Contract', 'Order', 'Quote', 
    'RecordType', 'Profile', 'PermissionSet', 'Group', 'QueueSobject', 
    'EmailMessage', 'Task', 'Event', 
    'Campaign', 'CampaignMember', 'Solution', 'ContentDocument', 
    'CollaborationGroup', 'Idea', 'LiveChatTranscript', 'AuthSession'
]

API_VERSION = '65.0'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, f'Picklist_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
DEBUG_MODE = False


# ===========================================
# HELPER CLASSES 
# ===========================================

class FieldInfo:
    """Represents picklist field metadata"""
    def __init__(self, api_name: str, label: str):
        self.api_name = api_name
        self.label = label


class PicklistValueDetail:
    """Represents a single picklist value"""
    def __init__(self, label: str, value: str, is_active: bool = True):
        self.label = label
        self.value = value
        self.is_active = is_active


class ProcessingResult:
    """Stores processing results for an object"""
    def __init__(self):
        self.values_processed = 0
        self.inactive_values = 0
        self.rows: List[List[str]] = []
        self.picklist_fields_count = 0
        self.object_exists = True
        self.error_message = None


# ===========================================
# MAIN EXPORT CLASS 
# ===========================================

class PicklistExporter:
    """Main exporter class with enhanced statistics"""
    
    def __init__(self, username: str, password: str, security_token: str, domain: str = 'login', status_callback=None):
        """Initialize Salesforce connection"""
        self.status_callback = status_callback
        if self.status_callback:
            self.status_callback("Initializing Salesforce Connection...")
            
        try:
            self.sf = Salesforce(
                username=username,
                password=password,
                security_token=security_token,
                domain=domain
            )
            self.base_url = f"https://{self.sf.sf_instance}"
            self.session_id = self.sf.session_id
            self.headers = {
                'Authorization': f'Bearer {self.session_id}',
                'Content-Type': 'application/json'
            }
            if self.status_callback:
                self.status_callback(f"✅ Connected to: {self.base_url}")
        except Exception as e:
            if self.status_callback:
                self.status_callback(f"❌ Connection failed: {str(e)}")
            raise
    
    def _log_status(self, message: str):
        """Internal helper to send log messages back to the GUI"""
        if self.status_callback:
            # Setting verbose=True ensures it goes to the detailed output, not just the console summary
            self.status_callback(message, verbose=True) 

    def export_picklists(self, object_names: List[str], output_path: str) -> Tuple[str, Dict]:
        """
        Main export method with enhanced statistics tracking
        Returns: (output_path, statistics_dict)
        """
        self._log_status("=== Starting Picklist Export ===")
        self._log_status(f"Total objects to process: {len(object_names)}")
        
        # Enhanced statistics tracking
        stats = {
            'total_objects': len(object_names),
            'successful_objects': 0,
            'failed_objects': 0,
            'objects_not_found': 0,
            'objects_with_zero_picklists': 0,
            'objects_with_picklists': 0,
            'total_picklist_fields': 0,
            'total_values': 0,
            'total_active_values': 0,
            'total_inactive_values': 0,
            'failed_object_details': [],
            'objects_without_picklists': [],
            'objects_not_found_list': []
        }
        
        all_rows = []
        header = ['Object', 'Field Label', 'Field API', 'Picklist Value Label', 'Picklist Value API', 'Status']
        all_rows.append(header)
        
        # Process each object
        for i, obj_name in enumerate(object_names, 1):
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            
            try:
                result = self._process_object(obj_name)
                
                # Track different scenarios
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({'name': obj_name, 'reason': 'Object does not exist in org'})
                    self._log_status(f"  ⚠️  Object not found in org")
                    
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  ℹ️  No picklist fields found")
                    
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    all_rows.extend(result.rows)
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    
                    self._log_status(f"  ✅ Fields: {result.picklist_fields_count}, Active: {result.values_processed - result.inactive_values}, Inactive: {result.inactive_values}")
                
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({'name': obj_name, 'reason': error_msg})
                continue
            
            self._log_status("") # newline
        
        # Create Excel file
        self._log_status("=== Creating Excel File ===")
        final_output_path = self._create_excel_file(all_rows, output_path)
        
        return final_output_path, stats
    
    def _process_object(self, obj_name: str) -> ProcessingResult:
        """Process a single object"""
        result = ProcessingResult()
        
        # Check if object exists
        try:
            getattr(self.sf, obj_name).describe()
        except Exception as e:
            if 'NOT_FOUND' in str(e) or 'INVALID_TYPE' in str(e):
                result.object_exists = False
                return result
            raise
        
        # Get picklist fields
        picklist_fields = self._get_picklist_fields(obj_name)
        result.picklist_fields_count = len(picklist_fields)
        
        if not picklist_fields:
            return result
        
        self._log_status(f"  Found {len(picklist_fields)} picklist fields")
        
        # Resolve EntityDefinition Id
        entity_def_id = self._resolve_entity_definition_id(obj_name)
        if entity_def_id:
            self._log_status(f"  EntityDefinition.Id: {entity_def_id}")
        
        # Process each field
        for field_api, field_info in picklist_fields.items():
            values = self._query_picklist_values_with_fallback(
                obj_name, entity_def_id, field_api
            )
            
            if not values:
                self._log_status(f"    Field: {field_api} - No values found")
                continue
            
            self._log_status(f"    Field: {field_api} - {len(values)} values")
            
            # Build rows
            for value in values:
                is_active = value.is_active if value.is_active is not None else True
                status = 'Active' if is_active else 'Inactive'
                
                if not is_active:
                    result.inactive_values += 1
                
                row = [
                    obj_name,
                    field_info.label,
                    field_api,
                    value.label,
                    value.value,
                    status
                ]
                result.rows.append(row)
                result.values_processed += 1
        
        return result
    
    def _get_picklist_fields(self, object_name: str) -> Dict[str, FieldInfo]:
        """Get all picklist fields"""
        fields_dict = {}
        
        try:
            obj_describe = getattr(self.sf, object_name).describe()
            
            for field in obj_describe['fields']:
                field_type = field['type']
                
                if field_type in ['picklist', 'multipicklist']:
                    field_info = FieldInfo(
                        api_name=field['name'],
                        label=field['label']
                    )
                    fields_dict[field_info.api_name] = field_info
        
        except Exception as e:
            self._log_status(f"  ERROR in _get_picklist_fields: {str(e)}")
        
        return fields_dict
    
    def _resolve_entity_definition_id(self, object_name: str) -> Optional[str]:
        """Resolve EntityDefinition Id"""
        try:
            query = f"SELECT Id FROM EntityDefinition WHERE QualifiedApiName = '{object_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            params = {'q': query}
            
            response = requests.get(url, headers=self.headers, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                records = data.get('records', [])
                if records:
                    return records[0]['Id']
        
        except Exception as e:
            self._log_status(f"  ERROR resolveEntityDefinitionId: {str(e)}")
        
        return None
    
    def _query_picklist_values_with_fallback(
        self, 
        object_name: str, 
        entity_def_id: Optional[str], 
        field_name: str
    ) -> List[PicklistValueDetail]:
        """Try multiple approaches to get picklist values"""
        
        values = self._query_field_definition_tooling(object_name, field_name)
        if values:
            return values
        
        if entity_def_id:
            values = self._query_custom_field_tooling(entity_def_id, field_name)
            if values:
                return values
        
        values = self._query_custom_field_tooling_table_enum(object_name, field_name)
        if values:
            return values
        
        values = self._query_rest_describe_for_picklist(object_name, field_name)
        if values:
            return values
        
        return []
    
    def _query_field_definition_tooling(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Fallback method 1"""
        try:
            query = f"SELECT Metadata FROM FieldDefinition WHERE EntityDefinition.QualifiedApiName = '{object_name}' AND QualifiedApiName = '{field_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            params = {'q': query}
            
            response = requests.get(url, headers=self.headers, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                records = data.get('records', [])
                if records:
                    metadata = records[0].get('Metadata', {})
                    return self._parse_value_set(metadata)
        
        except Exception as e:
            self._log_status(f"      ERROR queryFieldDefinitionTooling: {str(e)}")
        
        return []
    
    def _query_custom_field_tooling(self, entity_def_id: str, field_name: str) -> List[PicklistValueDetail]:
        """Fallback method 2"""
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{entity_def_id}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            params = {'q': query}
            
            response = requests.get(url, headers=self.headers, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                records = data.get('records', [])
                if records:
                    metadata = records[0].get('Metadata', {})
                    return self._parse_value_set(metadata)
        
        except Exception as e:
            self._log_status(f"      ERROR queryCustomFieldTooling: {str(e)}")
        
        return []
    
    def _query_custom_field_tooling_table_enum(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Fallback method 3"""
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{object_name}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{API_VERSION}/tooling/query/"
            params = {'q': query}
            
            response = requests.get(url, headers=self.headers, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                records = data.get('records', [])
                if records:
                    metadata = records[0].get('Metadata', {})
                    return self._parse_value_set(metadata)
        
        except Exception as e:
            self._log_status(f"      ERROR queryCustomFieldToolingTableEnum: {str(e)}")
        
        return []
    
    def _query_rest_describe_for_picklist(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Fallback method 4"""
        try:
            url = f"{self.base_url}/services/data/v{API_VERSION}/sobjects/{object_name}/describe"
            response = requests.get(url, headers=self.headers, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                fields = data.get('fields', [])
                
                for field in fields:
                    if field['name'].lower() == field_name.lower():
                        picklist_values = field.get('picklistValues', [])
                        results = []
                        for pv in picklist_values:
                            active_value = pv.get('active')
                            is_active = True if active_value is None else bool(active_value)
                            
                            detail = PicklistValueDetail(
                                label=pv.get('label', ''),
                                value=pv.get('value', ''),
                                is_active=is_active
                            )
                            results.append(detail)
                        return results
        
        except Exception as e:
            self._log_status(f"      ERROR queryRestDescribeForPicklist: {str(e)}")
        
        return []
    
    def _parse_value_set(self, metadata: dict) -> List[PicklistValueDetail]:
        """Parse valueSet from metadata"""
        results = []
        
        try:
            value_set = metadata.get('valueSet', {})
            if not value_set:
                return results
            
            if 'valueSetDefinition' in value_set:
                vsd = value_set['valueSetDefinition']
                values = vsd.get('value', [])
            elif 'value' in value_set:
                values = value_set['value']
            else:
                return results
            
            for v in values:
                label = v.get('label', '')
                value_name = v.get('valueName') or v.get('value', '')
                
                is_active_raw = v.get('isActive')
                if is_active_raw is None:
                    is_active = True
                else:
                    is_active = bool(is_active_raw)
                
                detail = PicklistValueDetail(
                    label=label,
                    value=value_name,
                    is_active=is_active
                )
                results.append(detail)
        
        except Exception as e:
            self._log_status(f"      ERROR parseValueSet: {str(e)}")
        
        return results
    
    def _create_excel_file(self, rows: List[List[str]], output_path: str) -> str:
        """Create Excel file with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Picklist Export"
        
        for row in rows:
            ws.append(row)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = "A2"
        
        wb.save(output_path)
        self._log_status(f"✅ Excel file created: {output_path}")
        self._log_status(f"✅ Total data rows: {len(rows) - 1}")
        
        return output_path

# ===========================================
# GUI HELPER FUNCTIONS (Must be defined before PicklistExportGUI)
# ===========================================

def format_runtime(seconds: float) -> str:
    """Format runtime in HH:MM:SS format"""
    td = timedelta(seconds=int(seconds))
    hours, remainder = divmod(td.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def print_statistics(stats: Dict, runtime_formatted: str, output_file: str):
    """Prints comprehensive statistics to the console (as detailed output for user)"""
    print("\n" + "=" * 70)
    print("✅ EXPORT COMPLETED SUCCESSFULLY! (Statistics Detail)")
    print("=" * 70)
    print(f"Total Runtime: {runtime_formatted}")
    print(f"Total Objects in List:          {stats['total_objects']}")
    print(f"✅ Successfully Processed:       {stats['successful_objects']}")
    print(f"❌ Failed to Process:            {stats['failed_objects']}")
    print(f"⚠️  Objects Not Found in Org:    {stats['objects_not_found']}")
    print(f"Total Picklist Fields:          {stats['total_picklist_fields']}")
    print(f"Total Picklist Values:          {stats['total_values']}")
    print(f"✅ Active Values:                {stats['total_active_values']}")
    print(f"❌ Inactive Values:              {stats['total_inactive_values']}")
    print(f"Output File: {output_file}")
    if stats['failed_objects'] > 0:
        print("\n❌ FAILED OBJECTS (REASONS):")
        for detail in stats['failed_object_details']:
            print(f"   • {detail['name']}: {detail['reason']}")
    print("=" * 70)


class PicklistExportGUI(ctk.CTk):
    def __init__(self, objects_list):
        super().__init__()
        
        self.title("Salesforce Picklist Exporter")
        self.geometry("800x600")
        self.sf_exporter: Optional[PicklistExporter] = None
        self.all_objects = sorted(objects_list)
        
        # Grid layout configuration
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        # Frames for the two screens
        self.login_frame = ctk.CTkFrame(self)
        self.export_frame = ctk.CTkFrame(self)
        
        # Place frames
        self.login_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        self.export_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        
        self._setup_login_frame()
        self._setup_export_frame()
        
        self.export_frame.grid_forget() # Start with Screen 2 hidden

    # ==================================
    # Screen 1: Login & Authentication
    # ==================================
    
    def _setup_login_frame(self):
        login_frame = self.login_frame
        login_frame.columnconfigure(1, weight=1)
        
        ctk.CTkLabel(login_frame, text="Salesforce Login", font=ctk.CTkFont(size=24, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(20, 30))

        # Helper function for input rows
        def create_input_row(parent, row, label_text, password_mode=False):
            ctk.CTkLabel(parent, text=label_text, anchor="w").grid(row=row, column=0, padx=10, pady=10, sticky="w")
            entry = ctk.CTkEntry(parent, width=300, show="*" if password_mode else "")
            entry.grid(row=row, column=1, padx=10, pady=10, sticky="ew")
            return entry

        # Inputs
        self.username_entry = create_input_row(login_frame, 1, "Username:")
        self.password_entry = create_input_row(login_frame, 2, "Password:", password_mode=True)
        self.token_entry = create_input_row(login_frame, 3, "Security Token:", password_mode=True)

        # Org Type Radio Buttons
        ctk.CTkLabel(login_frame, text="Org Type:", anchor="w").grid(row=4, column=0, padx=10, pady=10, sticky="w")
        self.org_type_var = ctk.StringVar(value="Production")
        radio_prod = ctk.CTkRadioButton(login_frame, text="Production", variable=self.org_type_var, value="Production")
        radio_test = ctk.CTkRadioButton(login_frame, text="Sandbox/Test", variable=self.org_type_var, value="Sandbox")
        
        radio_prod.grid(row=4, column=1, padx=(10, 5), pady=10, sticky="w")
        radio_test.grid(row=4, column=1, padx=(140, 10), pady=10, sticky="w")
        
        # Login Button
        self.login_button = ctk.CTkButton(login_frame, text="Login to Salesforce", command=self.login_action, height=40)
        self.login_button.grid(row=5, column=0, columnspan=2, pady=30, sticky="ew", padx=10)

    def login_action(self):
        self.login_button.configure(state="disabled", text="Connecting...")
        
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        token = self.token_entry.get().strip()
        domain = 'test' if self.org_type_var.get() == 'Sandbox' else 'login'
        
        if not all([username, password, token]):
            messagebox.showerror("Input Error", "All fields (Username, Password, Security Token) are required.")
            self.login_button.configure(state="normal", text="Login to Salesforce")
            return

        try:
            # Instantiate PicklistExporter, passing the status callback
            self.sf_exporter = PicklistExporter(
                username=username, 
                password=password, 
                security_token=token, 
                domain=domain,
                status_callback=self.update_status
            )
            
            messagebox.showinfo("Success", "Successfully connected to Salesforce!")
            
            # Switch to Export Frame
            self.login_frame.grid_forget()
            self.export_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
            self.populate_object_list(self.all_objects)
            
        except Exception as e:
            messagebox.showerror("Login Failed", f"Connection Error: {str(e)}")
            self.sf_exporter = None
            self.login_button.configure(state="normal", text="Login to Salesforce")

    # ==================================
    # Screen 2: Object Selection & Export
    # ==================================
    
    def _setup_export_frame(self):
        export_frame = self.export_frame
        export_frame.grid_rowconfigure(2, weight=1)
        export_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(export_frame, text="Select Objects for Picklist Export", font=ctk.CTkFont(size=24, weight="bold")).grid(row=0, column=0, pady=(20, 10), sticky="n")

        # Search Bar
        search_frame = ctk.CTkFrame(export_frame)
        search_frame.grid(row=1, column=0, pady=10, padx=20, sticky="ew")
        search_frame.columnconfigure(1, weight=1)
        
        ctk.CTkLabel(search_frame, text="Search:").grid(row=0, column=0, padx=10, pady=10)
        self.search_entry = ctk.CTkEntry(search_frame, placeholder_text="Enter Object API Name...")
        self.search_entry.grid(row=0, column=1, padx=(0, 10), pady=10, sticky="ew")
        self.search_entry.bind("<KeyRelease>", self.filter_objects)

        # Object List Box (Multi-select) - Standard Tkinter widget inside CTkFrame
        self.object_listbox = tk.Listbox(export_frame, selectmode="multiple", height=15, exportselection=False,
                                         font=("Arial", 12), borderwidth=0, highlightthickness=0,
                                         selectbackground="#1F538D", fg="white", background="#242424") 
        self.object_listbox.grid(row=2, column=0, padx=20, pady=(0, 10), sticky="nsew")

        # Status Text Box (for detailed output)
        self.status_textbox = ctk.CTkTextbox(export_frame, height=100)
        self.status_textbox.grid(row=3, column=0, padx=20, pady=(10, 10), sticky="ew")
        self.status_textbox.insert("end", "Status: Ready to select objects and export.")
        self.status_textbox.configure(state="disabled")

        # Export Button
        self.export_button = ctk.CTkButton(export_frame, text="Export Picklist Data", command=self.export_action, height=40, fg_color="green")
        self.export_button.grid(row=4, column=0, pady=(10, 20), sticky="ew", padx=20)
    
    def update_status(self, message: str, verbose: bool = False):
        """Updates the GUI status text box with new messages"""
        timestamp = datetime.now().strftime("[%H:%M:%S]")
        display_message = f"{timestamp} {message}"
        
        self.status_textbox.configure(state="normal")
        self.status_textbox.insert("end", "\n" + display_message)
        self.status_textbox.see("end") # Auto-scroll to the bottom
        
        # Only print critical/summary messages to console for easy viewing
        if not verbose:
            print(display_message) 

        self.status_textbox.configure(state="disabled")
        self.update_idletasks() # Force GUI redraw for real-time update

    def populate_object_list(self, objects: List[str]):
        """Populates the Tkinter Listbox"""
        self.object_listbox.delete(0, END)
        for obj in objects:
            self.object_listbox.insert(END, obj)

    def filter_objects(self, event):
        """Filters the ListBox based on the search entry content"""
        search_term = self.search_entry.get().lower()
        filtered_objects = [
            obj for obj in self.all_objects 
            if search_term in obj.lower()
        ]
        self.populate_object_list(filtered_objects)
        
    def export_action(self):
        if not self.sf_exporter:
            messagebox.showerror("Error", "Not logged in. Please log in first.")
            return

        # 1. Get selected objects
        selected_indices = self.object_listbox.curselection()
        selected_objects = [self.object_listbox.get(i) for i in selected_indices]

        if not selected_objects:
            messagebox.showwarning("Warning", "Please select at least one object to export.")
            return

        # 2. Prompt for save location
        default_filename = f'Picklist_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not output_file_path:
            return # User cancelled

        self.export_button.configure(state="disabled", text="Exporting... DO NOT CLOSE")
        self.update_status(f"Starting export for {len(selected_objects)} objects to {output_file_path}...")
        start_time = time.time()
        
        # Initialize variables for the finally block
        output_path = None
        stats = None
        
        try:
            # 3. Call the core export logic
            output_path, stats = self.sf_exporter.export_picklists(
                selected_objects, 
                output_file_path
            )
            
            end_time = time.time()
            runtime_seconds = end_time - start_time
            runtime_formatted = format_runtime(runtime_seconds)
            
            self.update_status(f"Export Complete! Total Runtime: {runtime_formatted}")
            messagebox.showinfo("Export Done", f"Picklist data successfully exported to:\n{output_path}")
            
            # Print detailed statistics to the console for the user to copy/review
            print_statistics(stats, runtime_formatted, output_path)
            
        except Exception as e:
            self.update_status(f"❌ FATAL EXPORT ERROR: {str(e)}")
            messagebox.showerror("Export Error", f"A fatal error occurred during export: {str(e)}")
            
        finally:
            self.export_button.configure(state="normal", text="Export Picklist Data")

# ===========================================
# MAIN EXECUTION
# ===========================================

if __name__ == "__main__":
    try:
        app = PicklistExportGUI(OBJECTS_TO_EXPORT)
        app.mainloop()
    except Exception as e:
        print(f"\n❌ GUI Application Failed: {str(e)}")
        sys.exit(1)